"""
NEA Loss Analysis System - Views
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.views import View
from django.views.generic import ListView, DetailView
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum, Avg, Count, Q, Exists, OuterRef
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
import json
import uuid
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import datetime
import decimal

from .models import (
    NEAUser, LossReport, MonthlyLossData, MeterPoint, MeterReading,
    ConsumerCategory, EnergyUtilisation, ConsumerCount, FiscalYear,
    DistributionCenter, ProvincialOffice, Province, Notification, AuditLog,
    ProvincialReport, MonthlyMeterPointStatus, DCYearlyTarget, Message,
)


def home_redirect(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return redirect('login')


def logout_view(request):
    logout(request)
    return redirect('login')


# ─────────────────────────── AUTH VIEWS ───────────────────────────

class LoginView(View):
    template_name = 'nea_loss/login.html'

    def get(self, request):
        if request.user.is_authenticated:
            return redirect('dashboard')
        return render(request, self.template_name)

    def post(self, request):
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            user.last_login_ip = request.META.get('REMOTE_ADDR')
            user.save(update_fields=['last_login_ip'])
            AuditLog.objects.create(
                user=user, action='LOGIN', model_name='NEAUser',
                object_id=user.pk, description=f"User {user.username} logged in",
                ip_address=request.META.get('REMOTE_ADDR')
            )
            return redirect('dashboard')
        messages.error(request, 'Invalid username or password.')
        return render(request, self.template_name)


class ProfileView(LoginRequiredMixin, View):
    template_name = 'nea_loss/profile.html'

    def get(self, request):
        return render(request, self.template_name, {'user': request.user})

    def post(self, request):
        user = request.user
        user.full_name = request.POST.get('full_name', user.full_name)
        user.phone = request.POST.get('phone', user.phone)
        user.designation = request.POST.get('designation', user.designation)
        user.email = request.POST.get('email', user.email)

        new_password = request.POST.get('new_password')
        if new_password:
            old_password = request.POST.get('old_password')
            if user.check_password(old_password):
                user.set_password(new_password)
                messages.success(request, 'Password updated successfully.')
            else:
                messages.error(request, 'Old password is incorrect.')
                return render(request, self.template_name, {'user': user})

        user.save()
        messages.success(request, 'Profile updated successfully.')
        return redirect('profile')


# ─────────────────────────── DASHBOARD ───────────────────────────

class DashboardView(LoginRequiredMixin, View):
    template_name = 'nea_loss/dashboard.html'

    def get(self, request):
        user = request.user
        active_fy = FiscalYear.objects.filter(is_active=True).first()
        context = {'active_fy': active_fy}

        if getattr(user, 'is_system_admin', False):
            context.update(self._get_admin_context(active_fy))
        elif user.is_top_management:
            context.update(self._get_top_management_context(active_fy))
        elif user.is_provincial:
            context.update(self._get_provincial_context(user, active_fy))
        else:
            context.update(self._get_dc_context(user, active_fy))

        context['notifications'] = Notification.objects.filter(
            recipient=user, is_read=False
        ).order_by('-created_at')[:5]

        return render(request, self.template_name, context)

    def _get_admin_context(self, active_fy):
        # Reuse the top-management dashboard numbers, then overlay admin-specific controls.
        base = self._get_top_management_context(active_fy)

        reports = LossReport.objects.filter(fiscal_year=active_fy) if active_fy else LossReport.objects.none()
        total_received = reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0
        total_utilised = reports.aggregate(s=Sum('total_utilised_kwh'))['s'] or 0
        total_loss = total_received - total_utilised
        overall_loss_pct = (total_loss / total_received * 100) if total_received > 0 else 0

        # Ensure all required context variables are properly initialized
        admin_context = {
            'admin_total_users': NEAUser.objects.count(),
            'admin_reports_total': reports.count(),
            'admin_recent_audits': AuditLog.objects.select_related('user').order_by('-timestamp')[:12],
            'admin_quicklink_active_fy': active_fy.year_bs if active_fy else '',
            'admin_overall_loss_pct': round(overall_loss_pct, 2),
            'mgmtProvData': base.get('mgmtProvData', []),  # Use empty list for admin
            'prov_monthly_detail': base.get('prov_monthly_detail', {}),
            'monthly_trend': base.get('monthly_trend', []),
            'top_5_loss': base.get('top_5_loss', []),
            'bottom_5_loss': base.get('bottom_5_loss', []),
        }
        
        base.update(admin_context)
        return base

    def _get_top_management_context(self, active_fy):
        """MD/DMD/Director: interactive read-only dashboard — approved reports only."""
        MONTH_NAMES = {
            1:'Shrawan',2:'Bhadra',3:'Ashwin',4:'Kartik',
            5:'Mangsir',6:'Poush',7:'Magh',8:'Falgun',
            9:'Chaitra',10:'Baisakh',11:'Jestha',12:'Ashadh'
        }
        reports = LossReport.objects.filter(
            fiscal_year=active_fy, status='APPROVED'
        ).select_related('distribution_center','distribution_center__provincial_office')          if active_fy else LossReport.objects.none()

        total_received = reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0
        total_utilised = reports.aggregate(s=Sum('total_utilised_kwh'))['s'] or 0
        total_loss = total_received - total_utilised
        overall_loss_pct = round(float(total_loss) / float(total_received) * 100, 4) if total_received else 0

        # Province-wise breakdown
        prov_data = []
        if active_fy and reports.exists():
            for po in ProvincialOffice.objects.prefetch_related('distribution_centers').all():
                po_reports = reports.filter(distribution_center__provincial_office=po)
                po_recv = float(po_reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0)
                po_util = float(po_reports.aggregate(s=Sum('total_utilised_kwh'))['s'] or 0)
                po_loss = po_recv - po_util
                po_pct  = round(po_loss / po_recv * 100, 4) if po_recv else 0
                dc_count = po_reports.values('distribution_center').distinct().count()
                prov_data.append({
                    'name': po.name,
                    'loss_pct': po_pct,
                    'received': po_recv,
                    'loss_kwh': po_loss,
                    'dc_count': dc_count,
                    'approved_count': po_reports.count(),
                })

        # DC-level details for interactive table (all approved DCs)
        dc_table = []
        if active_fy and reports.exists():
            for r in reports.order_by('distribution_center__provincial_office__name',
                                       'distribution_center__name'):
                dc_table.append({
                    'dc_name':   r.distribution_center.name,
                    'dc_code':   r.distribution_center.code,
                    'province':  r.distribution_center.provincial_office.name,
                    'month':     r.get_month_display(),
                    'month_num': r.month,
                    'received':  float(r.total_received_kwh),
                    'utilised':  float(r.total_utilised_kwh),
                    'loss_kwh':  float(r.total_loss_kwh),
                    'loss_pct':  round(float(r.cumulative_loss_percent) * 100, 4),
                    'status':    r.status,
                    'report_pk': r.pk,
                })

        # Month-wise aggregated trend (across all DCs)
        monthly_trend = {}
        if active_fy and reports.exists():
            all_monthly = MonthlyLossData.objects.filter(
                report__in=reports
            ).values('month','month_name').annotate(
                tot_recv=Sum('net_energy_received'),
                tot_util=Sum('total_energy_utilised'),
                tot_loss=Sum('loss_unit'),
            ).order_by('month')
            for row in all_monthly:
                recv = float(row['tot_recv'] or 0)
                loss = float(row['tot_loss'] or 0)
                monthly_trend[row['month_name']] = {
                    'received': recv,
                    'utilised': float(row['tot_util'] or 0),
                    'loss': loss,
                    'loss_pct': round(loss/recv*100,4) if recv else 0,
                }

        # Top/bottom 5 DCs by loss %
        sorted_dc = sorted(dc_table, key=lambda x: x['loss_pct'], reverse=True)
        top_5_loss    = sorted_dc[:5]
        bottom_5_loss = sorted_dc[-5:] if len(sorted_dc) >= 5 else sorted_dc

        # Province list for filter dropdown
        all_provinces = list(ProvincialOffice.objects.values_list('name', flat=True))

        # Build per-DC monthly detail for sidebar browsing
        # Each DC: list of months with received, utilised, loss_unit, monthly_loss_pct, cumul_loss_pct
        dc_monthly_detail = {}  # dc_name -> list of month dicts
        all_monthly_data = (
            MonthlyLossData.objects
            .filter(report__in=reports)
            .select_related('report__distribution_center','report__distribution_center__provincial_office')
            .order_by('report__distribution_center__name', 'month')
        )
        
        # Group by DC and calculate cumulative correctly
        from collections import defaultdict
        dc_data = defaultdict(list)  # dc_name -> list of monthly data
        for md in all_monthly_data:
            dc_name = md.report.distribution_center.name
            dc_data[dc_name].append(md)
        
        # Calculate cumulative loss for each DC
        for dc_name, monthly_data_list in dc_data.items():
            dc_code = monthly_data_list[0].report.distribution_center.code
            province = monthly_data_list[0].report.distribution_center.provincial_office.name
            report_pk = monthly_data_list[0].report.pk
            
            cumulative_received = 0
            cumulative_utilised = 0
            
            dc_monthly_detail[dc_name] = {
                'dc_name': dc_name,
                'dc_code': dc_code,
                'province': province,
                'report_pk': report_pk,
                'months': [],
            }
            
            # Sort by month and calculate progressive cumulative
            monthly_data_list.sort(key=lambda x: x.month)
            for md in monthly_data_list:
                recv = float(md.net_energy_received)
                utilised = float(md.total_energy_utilised)
                loss = float(md.loss_unit)
                mpct = round(loss / recv * 100, 4) if recv else 0
                
                # Add to cumulative
                cumulative_received += recv
                cumulative_utilised += utilised
                
                # Calculate cumulative loss %
                cumulative_loss = cumulative_received - cumulative_utilised
                cpct = round(cumulative_loss / cumulative_received * 100, 4) if cumulative_received else 0
                
                dc_monthly_detail[dc_name]['months'].append({
                    'month_name': md.month_name,
                    'month': md.month,
                    'received': recv,
                    'utilised': utilised,
                    'loss_unit': loss,
                    'monthly_loss_pct': mpct,
                    'cumul_loss_pct': cpct,  # Now using correct progressive calculation
                })

        # Build per-province monthly detail for sidebar browsing
        prov_monthly_detail = {}
        for po in ProvincialOffice.objects.all():
            po_monthly_qs = MonthlyLossData.objects.filter(
                report__in=reports,
                report__distribution_center__provincial_office=po,
            ).values('month','month_name').annotate(
                tot_recv=Sum('net_energy_received'),
                tot_util=Sum('total_energy_utilised'),
                tot_loss=Sum('loss_unit'),
            ).order_by('month')
            months_list = []
            cum_recv = 0.0
            cum_loss = 0.0
            for row in po_monthly_qs:
                recv = float(row['tot_recv'] or 0)
                loss = float(row['tot_loss'] or 0)
                cum_recv += recv
                cum_loss += loss
                months_list.append({
                    'month_name': row['month_name'],
                    'month': row['month'],
                    'received': recv,
                    'utilised': float(row['tot_util'] or 0),
                    'loss_unit': loss,
                    'monthly_loss_pct': round(loss/recv*100,4) if recv else 0,
                    'cumul_loss_pct': round(cum_loss/cum_recv*100,4) if cum_recv else 0,
                })
            if months_list:
                prov_monthly_detail[po.name] = months_list

        # Build dc_report_table for Report Explorer (for top management) with correct cumulative loss calculation
        from nea_loss.models import DCYearlyTarget
        from collections import defaultdict
        
        # Load all monthly data for approved reports
        monthly_qs = MonthlyLossData.objects.filter(
            report__in=reports
        ).select_related('report__distribution_center').order_by('report__distribution_center__name', 'month')

        # Group monthly data by DC
        dc_monthly = defaultdict(dict)  # dc_id -> {month -> MonthlyLossData}
        for md in monthly_qs:
            dc_monthly[md.report.distribution_center_id][md.month] = md

        # Load provincial yearly targets
        targets = {}
        if active_fy:
            for t in DCYearlyTarget.objects.filter(fiscal_year=active_fy):
                targets[t.distribution_center_id] = float(t.target_loss_percent)

        # Build table rows
        dc_report_table = []
        # Only show months that have approved reports
        approved_months = sorted(set(
            md.month for md in monthly_qs
        ))
        
        # If no approved months, show empty
        if not approved_months:
            approved_months = []

        for dc in DistributionCenter.objects.filter(is_active=True).order_by('name'):
            dc_report = reports.filter(distribution_center=dc).first()
            month_rows = []
            
            # Calculate cumulative loss progressively for each approved month
            cumulative_received = 0
            cumulative_utilised = 0
            
            # Get all months in order and calculate cumulative progressively
            for m in approved_months:
                md = dc_monthly.get(dc.pk, {}).get(m)
                target = targets.get(dc.pk)  # Use yearly target for all months
                
                # Always calculate cumulative, but only add if DC has data for this month
                if md and md.net_energy_received:
                    cumulative_received += float(md.net_energy_received)
                if md and md.total_energy_utilised:
                    cumulative_utilised += float(md.total_energy_utilised)
                
                # Calculate cumulative loss % up to this month
                cumulative_loss = cumulative_received - cumulative_utilised
                cumulative_loss_pct = round(cumulative_loss / cumulative_received * 100, 4) if cumulative_received else 0
                
                # Only include month row if this DC has approved data for this month
                if md:  # Only show months where DC has data
                    month_rows.append({
                        'month': m,
                        'month_name': MONTH_NAMES.get(m, ''),
                        'received': float(md.net_energy_received) if md else None,
                        'sold': float(md.total_energy_utilised) if md else None,
                        'loss_unit': float(md.loss_unit) if md else None,
                        'monthly_loss_pct': round(float(md.monthly_loss_percent) * 100, 4) if md else None,
                        'cumulative_loss_pct': cumulative_loss_pct,  # This will show proper cumulative
                        'target': target,
                        'status': 'APPROVED',  # All shown data is approved
                    })
            dc_report_table.append({
                'dc': dc,
                'report': dc_report,
                'month_rows': month_rows,
                'has_data': any(r['received'] is not None for r in month_rows),
            })

        return {
            'total_received_kwh':   float(total_received),
            'total_utilised_kwh':   float(total_utilised),
            'total_loss_kwh':       float(total_loss),
            'overall_loss_pct':     overall_loss_pct,
            'total_dc_count':       DistributionCenter.objects.filter(is_active=True).count(),
            'reports_approved':     reports.count(),
            'reports_submitted':    LossReport.objects.filter(
                                        fiscal_year=active_fy,
                                        status__in=['SUBMITTED','PROVINCIAL_REVIEWED']
                                    ).count() if active_fy else 0,
            'reports_pending':      LossReport.objects.filter(
                                        fiscal_year=active_fy, status='DRAFT'
                                    ).count() if active_fy else 0,
            'dc_table':             dc_table,
            'provincial_data':      prov_data,
            'monthly_trend':        monthly_trend,
            'top_5_loss':           top_5_loss,
            'bottom_5_loss':        bottom_5_loss,
            'nea_target_pct':       float(active_fy.loss_target_percent) if active_fy else 3.35,
            'target_loss_pct':      float(active_fy.loss_target_percent) if active_fy else 3.35,
            'all_provinces':        all_provinces,
            'month_names_list':     list(MONTH_NAMES.values()),
            'dc_monthly_detail':    dc_monthly_detail,      # for sidebar DC browser
            'prov_monthly_detail':  prov_monthly_detail,    # for sidebar province browser
            'dc_report_table':      dc_report_table if 'dc_report_table' in locals() else [],  # Safe: only if exists
        }

    def _get_provincial_context(self, user, active_fy):
        po = user.provincial_office
        MONTH_NAMES = {
            1:'Shrawan',2:'Bhadra',3:'Ashwin',4:'Kartik',
            5:'Mangsir',6:'Poush',7:'Magh',8:'Falgun',
            9:'Chaitra',10:'Baisakh',11:'Jestha',12:'Ashadh'
        }
        all_reports = LossReport.objects.filter(
            fiscal_year=active_fy,
            distribution_center__provincial_office=po
        ).select_related('distribution_center') if active_fy else LossReport.objects.none()

        approved_reports = all_reports.filter(status='APPROVED')

        # Build month-wise summary table per DC (Excel format):
        # For each DC: one row per month with received, sold, loss units, monthly loss %, cumulative loss %
        # Plus provincial target per month if set
        dcs = DistributionCenter.objects.filter(provincial_office=po).order_by('name')

        # Load all monthly data for approved reports under this province
        from nea_loss.models import MonthlyLossData, DCYearlyTarget
        monthly_qs = MonthlyLossData.objects.filter(
            report__in=approved_reports
        ).select_related('report__distribution_center').order_by('report__distribution_center__name', 'month')

        # Group monthly data by DC
        from collections import defaultdict
        dc_monthly = defaultdict(dict)  # dc_id -> {month -> MonthlyLossData}
        for md in monthly_qs:
            dc_monthly[md.report.distribution_center_id][md.month] = md

        # Load provincial yearly targets
        targets = {}
        if active_fy:
            for t in DCYearlyTarget.objects.filter(
                fiscal_year=active_fy,
                distribution_center__provincial_office=po
            ):
                targets[t.distribution_center_id] = float(t.target_loss_percent)

        # Build table rows
        dc_report_table = []
        # Only show months that have approved reports
        # Since monthly_qs already filters approved_reports, we just need to get unique months
        approved_months = sorted(set(
            md.month for md in monthly_qs
        ))
        
        # If no approved months, show empty
        if not approved_months:
            approved_months = []

        for dc in dcs:
            dc_report = all_reports.filter(distribution_center=dc).first()
            month_rows = []
            
            # Calculate cumulative loss progressively for each approved month
            cumulative_received = 0
            cumulative_utilised = 0
            
            # Get all months in order and calculate cumulative progressively
            for m in approved_months:
                md = dc_monthly.get(dc.pk, {}).get(m)
                target = targets.get(dc.pk)  # Use yearly target for all months
                
                # Always calculate cumulative, but only add if DC has data for this month
                if md and md.net_energy_received:
                    cumulative_received += float(md.net_energy_received)
                if md and md.total_energy_utilised:
                    cumulative_utilised += float(md.total_energy_utilised)
                
                # Calculate cumulative loss % up to this month
                cumulative_loss = cumulative_received - cumulative_utilised
                cumulative_loss_pct = round(cumulative_loss / cumulative_received * 100, 4) if cumulative_received else 0
                
                # Only include month row if this DC has approved data for this month
                if md:  # Only show months where DC has data
                    month_rows.append({
                        'month': m,
                        'month_name': MONTH_NAMES.get(m, ''),
                        'received': float(md.net_energy_received) if md else None,
                        'sold': float(md.total_energy_utilised) if md else None,
                        'loss_unit': float(md.loss_unit) if md else None,
                        'monthly_loss_pct': round(float(md.monthly_loss_percent) * 100, 4) if md else None,
                        'cumulative_loss_pct': cumulative_loss_pct,  # This will show proper cumulative
                        'target': target,
                        'status': 'APPROVED',  # All shown data is approved
                    })
            dc_report_table.append({
                'dc': dc,
                'report': dc_report,
                'month_rows': month_rows,
                'has_data': any(r['received'] is not None for r in month_rows),
            })

        # Totals across all approved reports
        total_received = approved_reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0
        total_utilised = approved_reports.aggregate(s=Sum('total_utilised_kwh'))['s'] or 0
        total_loss = float(total_received) - float(total_utilised)
        overall_loss_pct = round(total_loss / float(total_received) * 100, 4) if total_received else 0

        return {
            'provincial_office': po,
            'dc_count': dcs.count(),
            'reports_submitted': all_reports.filter(status__in=['SUBMITTED','PROVINCIAL_REVIEWED']).count(),
            'reports_approved': all_reports.filter(status='APPROVED').count(),
            'reports_rejected': all_reports.filter(status='REJECTED').count(),
            'pending_review': all_reports.filter(status__in=['SUBMITTED','PROVINCIAL_REVIEWED']).select_related('distribution_center').order_by('-updated_at')[:10],
            'total_received_kwh': float(total_received),
            'total_loss_kwh': total_loss,
            'overall_loss_pct': overall_loss_pct,
            'target_loss_pct': float(active_fy.loss_target_percent) if active_fy else 3.35,
            'dc_report_table': dc_report_table,
            'all_months': [MONTH_NAMES[m] for m in approved_months],
            'all_month_nums': approved_months,
        }

    def _get_dc_context(self, user, active_fy):
        dc = user.distribution_center
        report = LossReport.objects.filter(
            distribution_center=dc, fiscal_year=active_fy
        ).first() if (dc and active_fy) else None

        MONTH_NAMES = {
            1:'Shrawan', 2:'Bhadra', 3:'Ashwin', 4:'Kartik',
            5:'Mangsir', 6:'Poush', 7:'Magh', 8:'Falgun',
            9:'Chaitra', 10:'Baisakh', 11:'Jestha', 12:'Ashadh'
        }
        ALL_MONTHS = list(range(1, 13))

        # Provincial yearly targets for this DC
        prov_targets = {}
        if active_fy and dc:
            yearly_target = DCYearlyTarget.objects.filter(
                distribution_center=dc, fiscal_year=active_fy
            ).first()
            prov_targets = float(yearly_target.target_loss_percent) if yearly_target else None

        # Initialize variables
        dc_monthly_cols = []
        approved_months = []
        ytd_received = 0
        ytd_loss = 0
        ytd_sold = 0
        ytd_loss_pct = 0

        # Build columns from ALL approved reports for current DC
        approved_reports = LossReport.objects.filter(
            distribution_center=dc,
            fiscal_year=active_fy,
            status='APPROVED'
        ).order_by('created_at')
        
        if approved_reports:
            # Collect all monthly data from ALL approved reports
            all_monthly_data = []
            for report in approved_reports:
                all_monthly_data.extend(report.monthly_data.all())
            
            # Sort by month number to ensure proper order
            all_monthly_data.sort(key=lambda x: x.month)
            
            cumulative_received = decimal.Decimal('0')
            cumulative_loss = decimal.Decimal('0')
            
            # Build feeder readings for all approved reports
            month_pks = [md.pk for md in all_monthly_data]
            deleted_pairs = set(
                MonthlyMeterPointStatus.objects.filter(
                    monthly_data_id__in=month_pks, is_active=False
                ).values_list('monthly_data_id', 'meter_point_id')
            )
            all_readings = (
                MeterReading.objects
                .filter(monthly_data__in=all_monthly_data)
                .select_related('meter_point')
                .order_by('meter_point__source_type', 'meter_point__name')
            )
            reading_map = {(r.monthly_data_id, r.meter_point_id): r for r in all_readings}

            import_types = {'SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP', 'ENERGY_IMPORT'}
            export_types = {'EXPORT_DC', 'EXPORT_IPP', 'ENERGY_EXPORT'}
            all_feeders = list(
                MeterPoint.objects.filter(
                    distribution_center=dc,
                    is_active=True,
                    source_type__in=list(import_types | export_types)
                ).order_by('source_type', 'name')
            )

            # Build columns for ALL approved months from ALL reports
            for md in all_monthly_data:
                received   = float(md.net_energy_received)
                sold       = float(md.total_energy_utilised)
                import_kwh = float(md.total_energy_import)
                export_kwh = float(md.total_energy_export)
                loss_unit  = float(md.loss_unit)

                monthly_loss_pct = round(loss_unit / received * 100, 4) if received else 0
                cumulative_received += md.net_energy_received
                cumulative_loss     += md.loss_unit
                cum_loss_pct = round(
                    float(cumulative_loss) / float(cumulative_received) * 100, 4
                ) if cumulative_received else 0

                prov_target = prov_targets

                # Build feeder list for this month
                feeders = []
                for mp in all_feeders:
                    if (md.pk, mp.pk) in deleted_pairs:
                        continue
                    r = reading_map.get((md.pk, mp.pk))
                    feeders.append({
                        'name': mp.name,
                        'type': mp.get_source_type_display(),
                        'is_export': mp.source_type in export_types,
                        'prev': float(r.previous_reading) if r else None,
                        'pres': float(r.present_reading)  if r else None,
                        'mf':   float(r.multiplying_factor) if r else 1,
                        'kwh':  float(r.unit_kwh) if r else None,
                    })

                dc_monthly_cols.append({
                    'month':            md.month,
                    'month_name':       MONTH_NAMES[md.month],
                    'has_data':         True,
                    'import_kwh':       import_kwh,
                    'export_kwh':       export_kwh,
                    'received':         received,
                    'sold':             sold,
                    'loss_unit':        loss_unit,
                    'monthly_loss_pct': monthly_loss_pct,
                    'cumulative_loss_pct': cum_loss_pct,
                    'prov_target':      prov_target,
                    'status':           'APPROVED',
                    'report_pk':        md.report.pk,
                    'feeders':          feeders,
                })

            # YTD totals from all approved months
            ytd_received = float(cumulative_received)
            ytd_loss     = float(cumulative_loss)
            ytd_sold     = sum(c['sold'] for c in dc_monthly_cols if c['sold'] is not None)
            ytd_loss_pct = round(ytd_loss / ytd_received * 100, 4) if ytd_received else 0
            
            # Approved months are the columns we just built
            approved_months = dc_monthly_cols

        return {
            'distribution_center':    dc,
            'current_report':         report,
            'dc_monthly_cols':        dc_monthly_cols,
            'approved_months':        approved_months,
            'nea_target_pct':         float(active_fy.loss_target_percent) if active_fy else None,
            'ytd_received':           ytd_received,
            'ytd_sold':               ytd_sold,
            'ytd_loss':               ytd_loss,
            'ytd_loss_pct':           ytd_loss_pct,
            'prov_targets':           prov_targets,
            'can_create_loss_report': _can_create_loss_report(user),
        }


# ─────────────────────────── REPORT VIEWS ───────────────────────────

class ReportListView(LoginRequiredMixin, View):
    template_name = 'nea_loss/reports/list.html'

    def get(self, request):
        user = request.user
        reports = LossReport.objects.select_related(
            'distribution_center', 'distribution_center__provincial_office',
            'fiscal_year', 'created_by'
        ).order_by('-fiscal_year__year_ad_start', 'distribution_center__name')

        if user.is_dc_level and user.distribution_center:
            # DC users see all their own reports (including drafts)
            reports = reports.filter(distribution_center=user.distribution_center)
        elif user.is_provincial:
            if user.provincial_office:
                # Provincial sees only SUBMITTED and above (not DRAFT) from their DCs
                reports = reports.filter(
                    distribution_center__provincial_office=user.provincial_office,
                    status__in=['SUBMITTED', 'PROVINCIAL_REVIEWED', 'APPROVED', 'REJECTED']
                )
            else:
                reports = reports.none()
        elif user.is_top_management:
            # MD/DMD/Director sees only APPROVED reports
            reports = reports.filter(status='APPROVED')
        elif getattr(user, 'is_system_admin', False):
            # System admin sees everything
            pass

        # Filters
        fy_id = request.GET.get('fiscal_year')
        status = request.GET.get('status')
        dc_id = request.GET.get('dc')
        search = request.GET.get('search')
        month_filter = request.GET.get('month', '')

        if fy_id:
            reports = reports.filter(fiscal_year_id=fy_id)
        if status:
            reports = reports.filter(status=status)
        if dc_id and not user.is_dc_level:
            reports = reports.filter(distribution_center_id=dc_id)
        if search:
            reports = reports.filter(distribution_center__name__icontains=search)
        if month_filter:
            try:
                reports = reports.filter(month=int(month_filter))
            except (ValueError, TypeError):
                pass

        paginator = Paginator(reports, 20)
        page = paginator.get_page(request.GET.get('page', 1))

        # UI helper: show Edit action only when the report is editable.
        for r in page:
            setattr(r, 'can_edit', _can_edit_report(request.user, r))

        # DC-level users cannot filter by other DCs — hide the dropdown entirely
        if user.is_dc_level:
            visible_dcs = []  # No DC filter shown to DC users
        elif user.is_provincial and user.provincial_office:
            visible_dcs = DistributionCenter.objects.filter(
                provincial_office=user.provincial_office, is_active=True
            )
        else:
            visible_dcs = DistributionCenter.objects.filter(is_active=True)

        MONTH_CHOICES = [
            (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
            (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
            (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
        ]

        return render(request, self.template_name, {
            'reports': page,
            'fiscal_years': FiscalYear.objects.all().order_by('-year_ad_start'),
            'distribution_centers': visible_dcs,
            'status_choices': LossReport.STATUS_CHOICES,
            'month_choices': MONTH_CHOICES,
            'selected_fy': fy_id,
            'selected_status': status,
            'selected_month': month_filter,
        })


class ReportCreateView(LoginRequiredMixin, View):
    template_name = 'nea_loss/reports/create.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and not _can_create_loss_report(request.user):
            messages.error(
                request,
                'Only distribution center and provincial office users can create loss reports. '
                'Head office roles can view reports, analytics, and approvals from the menu.',
            )
            return redirect('report_list')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        user = request.user
        dcs = DistributionCenter.objects.all()
        if user.is_dc_level and user.distribution_center:
            dcs = dcs.filter(pk=user.distribution_center.pk)
        elif user.is_provincial and user.provincial_office:
            # Provincial users should not create reports - disable DC selection
            dcs = DistributionCenter.objects.none()
            messages.info(
                request,
                'Provincial office reports are generated automatically from DC reports under your province. '
                'You can review and approve/reject reports from the dashboard.'
            )
            return redirect('report_list')

        # Calculate available months based on existing reports and approval status
        all_months = [
            (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
            (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
            (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
        ]
        
        # Filter months to only show available options
        available_months = []
        active_fy = FiscalYear.objects.filter(is_active=True).first()
        
        for month_num, month_name in all_months:
            # Check if this month can be created
            can_create = True
            
            # For months other than Shrawan, check if previous month is approved
            if month_num > 1:
                previous_month = month_num - 1
                previous_report = LossReport.objects.filter(
                    distribution_center__in=dcs,
                    fiscal_year=active_fy,
                    month=previous_month,
                    status='APPROVED'
                ).first()
                
                if not previous_report:
                    can_create = False
            
            # Check if report already exists for this month
            if active_fy:
                existing_report = LossReport.objects.filter(
                    distribution_center__in=dcs,
                    fiscal_year=active_fy,
                    month=month_num
                ).first()
                
                if existing_report:
                    can_create = False
            
            if can_create:
                available_months.append((month_num, month_name))
        
        months_list = available_months
        
        return render(request, self.template_name, {
            'fiscal_years': FiscalYear.objects.all(),
            'distribution_centers': dcs,
            'months_list': months_list,
        })

    def post(self, request):
        user = request.user
        fy_id = request.POST.get('fiscal_year')
        dc_id = request.POST.get('distribution_center')
        month_id = request.POST.get('month')

        month_names = {
            1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
            5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
            9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
        }

        if not fy_id or not dc_id or not month_id:
            messages.error(request, 'Please select distribution center, fiscal year, and month.')
            return self.get(request)

        try:
            fy = FiscalYear.objects.get(pk=fy_id)
            dc = DistributionCenter.objects.get(pk=dc_id)
            month = int(month_id)
        except (FiscalYear.DoesNotExist, DistributionCenter.DoesNotExist):
            messages.error(request, 'Selected fiscal year or distribution center is invalid.')
            return self.get(request)
        except (ValueError, TypeError):
            messages.error(request, 'Selected month is invalid.')
            return self.get(request)

        if month < 1 or month > 12:
            messages.error(request, 'Selected month is invalid.')
            return self.get(request)

        # Check if previous month is approved (except for Shrawan)
        if month > 1:  # Not Shrawan
            previous_month = month - 1
            try:
                previous_report = LossReport.objects.get(
                    distribution_center=dc,
                    fiscal_year=fy,
                    month=previous_month
                )
                if previous_report.status != 'APPROVED':
                    messages.error(
                        request,
                        f'Previous month report ({month_names.get(previous_month, "")}) must be approved before creating {month_names.get(month, "")} report.'
                    )
                    return self.get(request)
            except LossReport.DoesNotExist:
                messages.error(
                    request,
                    f'Previous month report ({month_names.get(previous_month, "")}) doesn\'t exist. Please create that first.'
                )
                return self.get(request)

        allowed = DistributionCenter.objects.all()
        if user.is_dc_level and user.distribution_center:
            allowed = allowed.filter(pk=user.distribution_center.pk)
        elif user.is_provincial and user.provincial_office:
            allowed = allowed.filter(provincial_office=user.provincial_office)
        else:
            allowed = allowed.none()
        if not allowed.filter(pk=dc.pk).exists():
            messages.error(request, 'You cannot create a report for this distribution center.')
            return redirect('report_list')

        # If this month report already exists, open the data-entry directly.
        existing_report = LossReport.objects.filter(
            fiscal_year=fy,
            distribution_center=dc,
            month=month,
        ).first()
        if existing_report:
            month_name = month_names.get(month, '')
            messages.info(request, f'Report for {month_name} already exists. Opening data entry...')
            return redirect('monthly_data', existing_report.pk, month)

        # Create new monthly report
        report = LossReport.objects.create(
            fiscal_year=fy,
            distribution_center=dc,
            month=month,
            created_by=request.user,
            status='DRAFT'
        )
        AuditLog.objects.create(
            user=request.user,
            action='CREATE',
            model_name='LossReport',
            object_id=report.pk,
            description=f"Created monthly loss report for {dc.name} - {fy.year_bs} - {month_names.get(month, '')}",
        )
        messages.success(request, f'Monthly report created for {dc.name}.')
        return redirect('monthly_data', report.pk, month)


class ReportDetailView(LoginRequiredMixin, View):
    template_name = 'nea_loss/reports/report_detail.html'

    def get(self, request, pk):
        report = get_object_or_404(LossReport, pk=pk)
        if not _can_view_report(request.user, report):
            messages.error(request, 'You do not have permission to view this report.')
            return redirect('report_list')
        # Only get monthly data if it exists (don't auto-create)
        try:
            monthly_data = MonthlyLossData.objects.filter(report=report).order_by('month')
        except:
            monthly_data = MonthlyLossData.objects.none()
        fy = report.fiscal_year
        months_entered = monthly_data.count()

        # Monthly chart data
        chart_data = {
            'months': [m.month_name for m in monthly_data],
            'received': [float(m.net_energy_received) for m in monthly_data],
            'utilised': [float(m.total_energy_utilised) for m in monthly_data],
            'loss_pct': [round(float(m.monthly_loss_percent) * 100, 2) for m in monthly_data],
            'cumulative_pct': [round(float(m.cumulative_loss_percent) * 100, 2) for m in monthly_data],
        }

        dc = request.user.distribution_center if request.user.is_dc_level else None
        start_month = dc.report_start_month if dc else 1
        all_months = [
            (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
            (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
            (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
        ]
        months_list = [(n, name) for n, name in all_months if n >= start_month]
        entered_months = [m.month for m in monthly_data]
        return render(request, self.template_name, {
            'report': report,
            'monthly_data': monthly_data,
            'months_entered': months_entered,
            'months_list': months_list,
            'entered_months': entered_months,
            'chart_data_json': json.dumps(chart_data),
            'target_pct': float(fy.loss_target_percent),
            'can_edit': _can_edit_report(request.user, report),
            'can_approve': _can_approve_report(request.user),
        })


class ReportEditView(LoginRequiredMixin, View):
    template_name = 'nea_loss/reports/edit.html'

    def get(self, request, pk):
        report = get_object_or_404(LossReport, pk=pk)
        if not _can_edit_report(request.user, report):
            messages.error(request, 'You do not have permission to edit this report.')
            return redirect('report_detail', pk=pk)
        return render(request, self.template_name, {'report': report})


class MonthlyDataView(LoginRequiredMixin, View):
    template_name = 'nea_loss/reports/monthly_data.html'

    def get(self, request, report_pk, month):
        report = get_object_or_404(LossReport, pk=report_pk)
        
        # Check if report is submitted/approved - normally view-only.
        # System admin and top management can edit APPROVED reports to make corrections.
        if _can_edit_report(request.user, report):
            can_edit = True
        elif _can_view_report(request.user, report):
            can_edit = False
        else:
            messages.error(request, 'You do not have permission to view this report.')
            return redirect('report_list')
            
        month_names = {
            1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
            5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
            9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
        }
        
        # Get previous month's present readings for each meter point
        previous_readings_dict = {}
        
        if month > 1:  # For months after Shrawan
            previous_month = month - 1
            try:
                # First check if previous month report exists and is approved
                previous_loss_report = LossReport.objects.get(
                    distribution_center=report.distribution_center,
                    fiscal_year=report.fiscal_year,
                    month=previous_month
                )
                
                if previous_loss_report.status != 'APPROVED':
                    messages.error(
                        request,
                        f'Previous month report ({month_names.get(previous_month, "")}) must be approved before creating {month_names.get(month, "")} report.'
                    )
                    return redirect('report_create')
                
                # Then get monthly data and create previous readings dictionary
                try:
                    previous_month_report = MonthlyLossData.objects.get(
                        report=previous_loss_report,
                        month=previous_month
                    )
                    # Create dictionary of previous readings for each meter point
                    # Only carry forward for non-single-reading types
                    for prev_reading in previous_month_report.meter_readings.select_related('meter_point').all():
                        if not prev_reading.meter_point.is_single_reading:
                            previous_readings_dict[prev_reading.meter_point_id] = prev_reading.present_reading or 0
                    
                    # Debug: Print what we found
                                
                except MonthlyLossData.DoesNotExist:
                    # Previous month report exists but no monthly data yet
                    # This is OK - just use 0 for all meter points
                    pass
                    
            except LossReport.DoesNotExist:
                # Previous month report doesn't exist
                if request.user.is_dc_level:  # Only show error to DC users
                    messages.error(
                        request,
                        f'Previous month report ({month_names.get(previous_month, "")}) hasn\'t been created. Please create that first.'
                    )
                    return redirect('report_create')
        else:
            # For Shrawan (month 1), previous reading is 0
            if month == 1:
                previous_month_present_reading = 0
        
        # Editable flows must have a MonthlyLossData row to persist AJAX saves.
        if can_edit:
            monthly, _ = MonthlyLossData.objects.get_or_create(
                report=report,
                month=month,
                defaults={'month_name': month_names.get(month, '')},
            )
        else:
            # View-only: do not create empty rows.
            monthly = MonthlyLossData.objects.filter(report=report, month=month).first()

        existing_readings = {r.meter_point_id: r for r in monthly.meter_readings.all()} if monthly else {}

        # IDs explicitly marked inactive for this month (deleted/disabled for this month only)
        inactive_for_month = set()
        if monthly:
            inactive_for_month = set(
                MonthlyMeterPointStatus.objects.filter(
                    monthly_data=monthly,
                    is_active=False
                ).values_list('meter_point_id', flat=True)
            )

        # Determine which meter points to show based on report status
        if report.status in ['APPROVED']:
            # Approved report: show feeders that had readings, BUT still respect
            # per-month soft-deletes — if a feeder was deleted for THIS specific month,
            # hide it from this month's view (other months are unaffected).
            meter_points = MeterPoint.objects.filter(
                distribution_center=report.distribution_center,
                is_active=True
            ).exclude(
                pk__in=inactive_for_month
            ).order_by('source_type', 'name')
        elif not monthly:
            # No monthly data yet — show all active meter points
            meter_points = MeterPoint.objects.filter(
                distribution_center=report.distribution_center,
                is_active=True
            ).order_by('source_type', 'name')
        else:
            # Draft/Rejected report: show all active feeders EXCEPT those
            # explicitly removed for this month via MonthlyMeterPointStatus
            meter_points = MeterPoint.objects.filter(
                distribution_center=report.distribution_center,
                is_active=True
            ).exclude(
                pk__in=inactive_for_month
            ).order_by('source_type', 'name')

        existing_readings = {r.meter_point_id: r for r in monthly.meter_readings.select_related('meter_point').all()} if monthly else {}

        # Identify new meter points (those created in the current month)
        # A meter point is "new for this month" ONLY if it has NO reading in any
        # prior month for this DC (approved or same-report draft).
        # If it has prior readings it MUST auto-fill previous reading from last month.
        all_meter_point_ids = set(meter_points.values_list('pk', flat=True))

        if month > 1:
            # IDs that already have at least one reading in a prior approved report
            prior_approved_ids = set(
                MeterReading.objects.filter(
                    meter_point__distribution_center=report.distribution_center,
                    monthly_data__report__status='APPROVED',
                    monthly_data__report__fiscal_year=report.fiscal_year,
                    monthly_data__month__lt=month,
                ).values_list('meter_point_id', flat=True).distinct()
            )
            # IDs that have a reading in a prior month of the same (possibly draft) report
            prior_same_report_ids = set(
                MeterReading.objects.filter(
                    monthly_data__report=report,
                    monthly_data__month__lt=month,
                ).values_list('meter_point_id', flat=True).distinct()
            )
            has_prior = prior_approved_ids | prior_same_report_ids
            # "New" = never appeared in any prior month → allow manual previous-reading entry
            new_meter_point_ids = all_meter_point_ids - has_prior
        else:
            # Shrawan (month 1): all feeders are first-time, previous reading is editable
            new_meter_point_ids = all_meter_point_ids

        import_points = meter_points.filter(
            source_type__in=['SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP', 'ENERGY_IMPORT']
        )
        export_points = meter_points.filter(source_type__in=['EXPORT_DC', 'EXPORT_IPP', 'ENERGY_EXPORT'])

        # IDs of single-reading meter points (ENERGY_IMPORT / ENERGY_EXPORT)
        single_reading_ids = set(
            meter_points.filter(source_type__in=['ENERGY_IMPORT', 'ENERGY_EXPORT']).values_list('pk', flat=True)
        )

        consumer_categories = ConsumerCategory.objects.filter(is_active=True).filter(
            Q(distribution_center__isnull=True)
            | Q(distribution_center_id=report.distribution_center_id)
        ).order_by('display_order', 'name')

        # Load DC yearly target for this fiscal year (set by provincial office)
        dc_yearly_target = None
        if report.fiscal_year:
            t = DCYearlyTarget.objects.filter(
                distribution_center=report.distribution_center,
                fiscal_year=report.fiscal_year,
            ).first()
            dc_yearly_target = float(t.target_loss_percent) if t else None

        existing_utilisations = {e.consumer_category_id: e for e in monthly.energy_utilisations.all()} if monthly else {}
        existing_counts = {c.consumer_category_id: c for c in monthly.consumer_counts.all()} if monthly else {}

        import_type_choices = [
            (k, v) for k, v in MeterPoint.SOURCE_TYPE_CHOICES
            if k in ['SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP', 'ENERGY_IMPORT']
        ]
        export_type_choices = [
            (k, v) for k, v in MeterPoint.SOURCE_TYPE_CHOICES
            if k in ['EXPORT_DC', 'EXPORT_IPP', 'ENERGY_EXPORT']
        ]

        months_nav = list(month_names.items())
        
        return render(request, self.template_name, {
            'report': report,
            'monthly': monthly,
            'month': month,  # Use URL parameter month instead of report.month
            'month_name': month_names.get(month, ''),
            'previous_readings_dict': previous_readings_dict,
            'months_nav': months_nav,
            'import_points': import_points,
            'export_points': export_points,
            'existing_readings': existing_readings,
            'consumer_categories': consumer_categories,
            'existing_utilisations': existing_utilisations,
            'existing_counts': existing_counts,
            'can_edit': can_edit,
            'import_type_choices': import_type_choices,
            'export_type_choices': export_type_choices,
            'new_meter_point_ids': list(new_meter_point_ids),
            'single_reading_ids': list(single_reading_ids),  # ENERGY_IMPORT / ENERGY_EXPORT
            'dc_yearly_target': dc_yearly_target,
        })


class ReportPrintView(LoginRequiredMixin, View):
    template_name = 'nea_loss/reports/print.html'

    def get(self, request, pk):
        report = get_object_or_404(LossReport, pk=pk)
        monthly_data = report.monthly_data.order_by('month').prefetch_related(
            'meter_readings__meter_point',
            'energy_utilisations__consumer_category',
            'consumer_counts__consumer_category'
        )
        return render(request, self.template_name, {
            'report': report,
            'monthly_data': monthly_data,
        })


# ─────────────────────────── REPORT ACTIONS ───────────────────────────

@login_required
def report_submit(request, pk):
    report = get_object_or_404(LossReport, pk=pk)
    if request.method == 'POST' and _can_edit_report(request.user, report):
        report.status = 'SUBMITTED'
        report.submitted_by = request.user
        report.submission_date = timezone.now()
        dc_remarks = (request.POST.get('remarks') or '').strip()
        if dc_remarks:
            report.remarks = dc_remarks
        report.save()

        # Notify provincial manager
        if report.distribution_center.provincial_office:
            prov_managers = NEAUser.objects.filter(
                role='PROVINCIAL_MANAGER',
                provincial_office=report.distribution_center.provincial_office
            )
            for mgr in prov_managers:
                Notification.objects.create(
                    recipient=mgr,
                    notification_type='REPORT_SUBMITTED',
                    title=f'New Report: {report.distribution_center.name}',
                    message=f'Loss report for {report.fiscal_year.year_bs} has been submitted by {report.distribution_center.name}.',
                    related_report=report
                )
        # No flash message: user feedback is shown in the UI flow.
    return redirect('report_detail', pk=pk)


@login_required
def report_approve(request, pk):
    report = get_object_or_404(LossReport, pk=pk)
    if request.method == 'POST' and _can_approve_report(request.user):
        report.status = 'APPROVED'
        report.approved_by = request.user
        report.approval_date = timezone.now()
        approval_remarks = (request.POST.get('remarks') or '').strip()
        if approval_remarks:
            report.remarks = approval_remarks
        report.save()
        messages.success(request, 'Report approved successfully.')
    return redirect('report_detail', pk=pk)


@login_required
def report_reject(request, pk):
    report = get_object_or_404(LossReport, pk=pk)
    if request.method == 'POST' and _can_approve_report(request.user):
        report.status = 'REJECTED'
        report.remarks = request.POST.get('remarks', '')
        report.save()
        messages.warning(request, 'Report has been rejected.')
    return redirect('report_detail', pk=pk)


@login_required
@require_POST
def report_delete(request, pk):
    report = get_object_or_404(LossReport, pk=pk)
    if not _can_edit_report(request.user, report):
        messages.error(request, 'You do not have permission to delete this report.')
        return redirect('report_detail', pk=pk)
    if report.status != 'DRAFT':
        messages.error(request, 'Only DRAFT reports can be deleted.')
        return redirect('report_detail', pk=pk)
    report.delete()
    messages.success(request, 'Report deleted successfully.')
    return redirect('report_list')


@login_required
@require_POST
def monthly_data_delete(request, report_pk, month):
    report = get_object_or_404(LossReport, pk=report_pk)
    if not _can_edit_report(request.user, report):
        messages.error(request, 'You do not have permission to delete this month data.')
        return redirect('report_detail', pk=report.pk)

    # Since each report is now monthly, delete the entire report
    month_names = {
        1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
        5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
        9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
    }
    month_name = month_names.get(month, '')
    
    # Delete the entire monthly report
    report.delete()
    messages.success(request, f'Deleted monthly report for {month_name}.')
    return redirect('report_list')


@login_required
@require_POST
def api_delete_meter_reading_for_month(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    monthly_id = data.get('monthly_id')
    meter_point_id = data.get('meter_point_id')
    if not monthly_id or not meter_point_id:
        return JsonResponse({'error': 'monthly_id and meter_point_id are required'}, status=400)

    monthly = get_object_or_404(MonthlyLossData, pk=monthly_id)
    if not _can_edit_report(request.user, monthly.report):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    MeterReading.objects.filter(monthly_data=monthly, meter_point_id=meter_point_id).delete()

    # Recalculate totals from remaining readings
    import_types = {'SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP'}
    export_types = {'EXPORT_DC', 'EXPORT_IPP'}
    total_import = decimal.Decimal('0')
    total_export = decimal.Decimal('0')
    for mr in monthly.meter_readings.select_related('meter_point').all():
        if mr.meter_point.source_type in import_types:
            total_import += mr.unit_kwh
        elif mr.meter_point.source_type in export_types:
            total_export += mr.unit_kwh

    monthly.total_energy_import = total_import
    monthly.total_energy_export = total_export
    monthly.net_energy_received = total_import - total_export
    monthly.loss_unit = monthly.net_energy_received - monthly.total_energy_utilised
    if monthly.net_energy_received > 0:
        monthly.monthly_loss_percent = monthly.loss_unit / monthly.net_energy_received
    else:
        monthly.monthly_loss_percent = 0
    monthly.save()
    monthly.report.calculate_summary()

    return JsonResponse({
        'success': True,
        'total_import': float(total_import),
        'total_export': float(total_export),
        'net_received': float(monthly.net_energy_received),
    })


@login_required
@require_POST
def api_disable_meter_point_for_month(request):
    """Disable a meter point for a specific month while preserving historical data"""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    monthly_id = data.get('monthly_id')
    meter_point_id = data.get('meter_point_id')
    if not monthly_id or not meter_point_id:
        return JsonResponse({'error': 'monthly_id and meter_point_id are required'}, status=400)

    monthly = get_object_or_404(MonthlyLossData, pk=monthly_id)
    if not _can_edit_report(request.user, monthly.report):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    meter_point = get_object_or_404(MeterPoint, pk=meter_point_id)
    if meter_point.distribution_center_id != monthly.report.distribution_center_id:
        return JsonResponse({'error': 'Invalid meter point for this distribution center'}, status=400)

    # Create or update the monthly status to mark this meter point as inactive for this month
    status, created = MonthlyMeterPointStatus.objects.update_or_create(
        monthly_data=monthly,
        meter_point=meter_point,
        defaults={'is_active': False}
    )

    # Delete the meter reading for this month (if it exists)
    MeterReading.objects.filter(monthly_data=monthly, meter_point=meter_point).delete()

    # Recalculate totals
    import_types = {'SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP'}
    export_types = {'EXPORT_DC', 'EXPORT_IPP'}
    total_import = decimal.Decimal('0')
    total_export = decimal.Decimal('0')
    for mr in monthly.meter_readings.select_related('meter_point').all():
        if mr.meter_point.source_type in import_types:
            total_import += mr.unit_kwh
        elif mr.meter_point.source_type in export_types:
            total_export += mr.unit_kwh

    monthly.total_energy_import = total_import
    monthly.total_energy_export = total_export
    monthly.net_energy_received = total_import - total_export
    monthly.loss_unit = monthly.net_energy_received - monthly.total_energy_utilised
    if monthly.net_energy_received > 0:
        monthly.monthly_loss_percent = monthly.loss_unit / monthly.net_energy_received
    else:
        monthly.monthly_loss_percent = 0
    monthly.save()
    monthly.report.calculate_summary()

    return JsonResponse({
        'success': True,
        'message': f'Meter point "{meter_point.name}" disabled for {monthly.month_name}. Historical data preserved.',
        'total_import': float(total_import),
        'total_export': float(total_export),
        'net_received': float(monthly.net_energy_received),
    })


@login_required
def report_export_excel(request, pk):
    report = get_object_or_404(LossReport, pk=pk)
    wb = _generate_excel_report(report)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="LossReport_{report.distribution_center.code}_{report.fiscal_year.year_bs.replace("/","_")}.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────── ORGANIZATION VIEWS ───────────────────────────

class OrgOverviewView(LoginRequiredMixin, View):
    template_name = 'nea_loss/organizations/overview.html'

    def get(self, request):
        provinces = Province.objects.prefetch_related(
            'offices__distribution_centers'
        ).all()
        return render(request, self.template_name, {'provinces': provinces})


class DCDetailView(LoginRequiredMixin, View):
    template_name = 'nea_loss/organizations/dc_detail.html'

    def get(self, request, pk):
        dc = get_object_or_404(DistributionCenter, pk=pk)
        reports = LossReport.objects.filter(distribution_center=dc).order_by('-fiscal_year__year_ad_start')
        meter_points = dc.meter_points.all()
        return render(request, self.template_name, {
            'dc': dc,
            'reports': reports,
            'meter_points': meter_points,
        })


# ─────────────────────────── ANALYTICS ───────────────────────────

class AnalyticsView(LoginRequiredMixin, View):
    template_name = 'nea_loss/analytics/overview.html'

    def get(self, request):
        user = request.user
        active_fy = FiscalYear.objects.filter(is_active=True).first()

        # Base queryset — only APPROVED reports feed analytics for all roles above DC
        if getattr(user, 'is_system_admin', False):
            # Sysadmin sees approved reports across all
            all_reports = LossReport.objects.filter(
                fiscal_year=active_fy, status='APPROVED'
            ) if active_fy else LossReport.objects.none()
        elif user.is_top_management:
            all_reports = LossReport.objects.filter(
                fiscal_year=active_fy, status='APPROVED'
            ) if active_fy else LossReport.objects.none()
        elif user.is_provincial and user.provincial_office:
            all_reports = LossReport.objects.filter(
                fiscal_year=active_fy,
                status='APPROVED',
                distribution_center__provincial_office=user.provincial_office,
            ) if active_fy else LossReport.objects.none()
        elif user.is_dc_level and user.distribution_center:
            # DC users see their own submitted+ reports in analytics
            all_reports = LossReport.objects.filter(
                fiscal_year=active_fy,
                distribution_center=user.distribution_center,
                status__in=['SUBMITTED', 'PROVINCIAL_REVIEWED', 'APPROVED'],
            ) if active_fy else LossReport.objects.none()
        else:
            all_reports = LossReport.objects.none()

        view_mode = request.GET.get('view', 'dc')  # 'dc' or 'province'

        # Loss by DC
        dc_data = all_reports.values(
            'distribution_center__name'
        ).annotate(
            total_received=Sum('total_received_kwh'),
            total_loss=Sum('total_loss_kwh')
        ).order_by('-total_loss')[:15]

        # Monthly trend
        monthly_trend = {}
        for report in all_reports:
            for md in report.monthly_data.all():
                m = md.month_name
                if m not in monthly_trend:
                    monthly_trend[m] = {'received': 0, 'loss': 0}
                monthly_trend[m]['received'] += float(md.net_energy_received)
                monthly_trend[m]['loss'] += float(md.loss_unit)

        # Province-wise aggregation
        prov_data = []
        for po in ProvincialOffice.objects.all():
            po_reports = all_reports.filter(distribution_center__provincial_office=po)
            po_received = po_reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0
            po_utilised = po_reports.aggregate(s=Sum('total_utilised_kwh'))['s'] or 0
            po_loss = float(po_received) - float(po_utilised)
            po_loss_pct = (po_loss / float(po_received) * 100) if po_received else 0
            prov_data.append({
                'name': po.name,
                'loss_pct': round(po_loss_pct, 2),
                'total_received': float(po_received),
                'total_loss': po_loss,
            })

        return render(request, self.template_name, {
            'active_fy': active_fy,
            'dc_data': list(dc_data),
            'monthly_trend': monthly_trend,
            'prov_data': prov_data,
            'view_mode': view_mode,
            'total_received': all_reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0,
            'total_loss': all_reports.aggregate(s=Sum('total_loss_kwh'))['s'] or 0,
            'target_pct': float(active_fy.loss_target_percent) if active_fy else 3.35,
        })


class ComparisonView(LoginRequiredMixin, View):
    template_name = 'nea_loss/analytics/comparison.html'

    def get(self, request):
        user = request.user
        fiscal_years = FiscalYear.objects.all().order_by('-year_ad_start')

        MONTH_NAMES = [
            'Shrawan','Bhadra','Ashwin','Kartik','Mangsir','Poush',
            'Magh','Falgun','Chaitra','Baisakh','Jestha','Ashadh'
        ]

        # Build datasets: one per fiscal year — monthly average loss % from APPROVED reports
        datasets = []
        for fy in fiscal_years:
            if user.is_dc_level and user.distribution_center:
                fy_reports = LossReport.objects.filter(
                    fiscal_year=fy,
                    distribution_center=user.distribution_center,
                    status__in=['SUBMITTED', 'PROVINCIAL_REVIEWED', 'APPROVED'],
                )
            elif user.is_provincial and user.provincial_office:
                fy_reports = LossReport.objects.filter(
                    fiscal_year=fy,
                    status='APPROVED',
                    distribution_center__provincial_office=user.provincial_office,
                )
            else:
                fy_reports = LossReport.objects.filter(fiscal_year=fy, status='APPROVED')

            # Aggregate per month: sum of received and loss across all DCs in this FY
            monthly_data = {}
            for report in fy_reports:
                for md in report.monthly_data.all():
                    m = md.month
                    if m not in monthly_data:
                        monthly_data[m] = {'received': 0.0, 'loss': 0.0}
                    monthly_data[m]['received'] += float(md.net_energy_received)
                    monthly_data[m]['loss'] += float(md.loss_unit)

            # Build 12-slot array of loss %, None for months with no data
            loss_pct_by_month = []
            has_data = False
            for m_num in range(1, 13):
                if m_num in monthly_data and monthly_data[m_num]['received'] > 0:
                    pct = round(monthly_data[m_num]['loss'] / monthly_data[m_num]['received'] * 100, 4)
                    loss_pct_by_month.append(pct)
                    has_data = True
                else:
                    loss_pct_by_month.append(None)

            if has_data:
                datasets.append({
                    'fy': fy.year_bs,
                    'data': loss_pct_by_month,
                })

        return render(request, self.template_name, {
            'fiscal_years': fiscal_years,
            'datasets': datasets,
            'month_names': MONTH_NAMES,
            'has_data': bool(datasets),
        })


# ─────────────────────────── DC YEARLY TARGETS (PROVINCIAL) ───────────────────────────

class DCYearlyTargetView(LoginRequiredMixin, View):
    """Provincial manager sets yearly loss % targets for each DC under their office."""
    template_name = 'nea_loss/reports/dc_yearly_targets.html'

    def dispatch(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            if not (getattr(user, 'is_system_admin', False) or user.is_provincial or user.is_top_management):
                messages.error(request, 'Only provincial managers can set DC yearly targets.')
                return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        user = request.user
        active_fy = FiscalYear.objects.filter(is_active=True).first()

        if getattr(user, 'is_system_admin', False) or user.is_top_management:
            dcs = DistributionCenter.objects.select_related('provincial_office').all().order_by('provincial_office__name', 'name')
        else:
            po = user.provincial_office
            dcs = DistributionCenter.objects.filter(provincial_office=po).order_by('name') if po else DistributionCenter.objects.none()

        # Build current targets map: {dc_id: target_loss_percent}
        existing_targets = {}
        if active_fy:
            for t in DCYearlyTarget.objects.filter(fiscal_year=active_fy, distribution_center__in=dcs):
                existing_targets[t.distribution_center_id] = float(t.target_loss_percent)

        dc_rows = []
        for dc in dcs:
            # Determine if this DC's target can be edited
            target_exists = dc.pk in existing_targets
            
            # Provincial users can only edit if target doesn't exist yet
            # Admin users can always edit
            can_edit = False
            if getattr(user, 'is_system_admin', False) or user.is_top_management:
                can_edit = True  # Admin can always edit
            elif user.is_provincial:
                can_edit = not target_exists  # Provincial can edit only if no target exists yet
            
            dc_rows.append({
                'dc': dc,
                'target': existing_targets.get(dc.pk, ''),
                'can_edit': can_edit,
                'target_exists': target_exists,
            })

        return render(request, self.template_name, {
            'active_fy': active_fy,
            'dc_rows': dc_rows,
        })

    def post(self, request):
        user = request.user
        active_fy = FiscalYear.objects.filter(is_active=True).first()
        if not active_fy:
            messages.error(request, 'No active fiscal year found.')
            return redirect('dc_yearly_targets')

        if getattr(user, 'is_system_admin', False) or user.is_top_management:
            allowed_dcs = set(DistributionCenter.objects.values_list('pk', flat=True))
        else:
            po = user.provincial_office
            allowed_dcs = set(DistributionCenter.objects.filter(provincial_office=po).values_list('pk', flat=True)) if po else set()

        saved = 0
        for key, val in request.POST.items():
            # key format: target_<dc_id>
            if not key.startswith('target_'):
                continue
            parts = key.split('_')
            if len(parts) != 2:
                continue
            try:
                dc_id = int(parts[1])
                val = val.strip()
                if not val:
                    # Delete existing target if blank submitted (admin only)
                    if getattr(user, 'is_system_admin', False) or user.is_top_management:
                        DCYearlyTarget.objects.filter(
                            distribution_center_id=dc_id, fiscal_year=active_fy
                        ).delete()
                    continue
                target_pct = float(val)
            except (ValueError, TypeError):
                continue

            if dc_id not in allowed_dcs:
                continue

            # Check if target already exists
            existing_target = DCYearlyTarget.objects.filter(
                distribution_center_id=dc_id, fiscal_year=active_fy
            ).first()
            
            # Provincial users cannot edit existing targets
            if existing_target and user.is_provincial:
                continue  # Skip this DC, provincial user cannot edit existing target
            
            # Admin users can always edit, provincial users can only create new targets
            DCYearlyTarget.objects.update_or_create(
                distribution_center_id=dc_id,
                fiscal_year=active_fy,
                defaults={'target_loss_percent': target_pct, 'set_by': user},
            )
            saved += 1

        messages.success(request, f'Saved {saved} yearly target(s) successfully.')
        return redirect('dc_yearly_targets')


# ─────────────────────────── DC MONTHLY TARGETS (DEPRECATED) ───────────────────────────

class DCMonthlyTargetView(LoginRequiredMixin, View):
    """Provincial manager sets monthly loss % targets for each DC under their office."""
    template_name = 'nea_loss/reports/dc_monthly_targets.html'

    def dispatch(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            if not (getattr(user, 'is_system_admin', False) or user.is_provincial or user.is_top_management):
                messages.error(request, 'Only provincial managers can set DC monthly targets.')
                return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        user = request.user
        active_fy = FiscalYear.objects.filter(is_active=True).first()

        if getattr(user, 'is_system_admin', False) or user.is_top_management:
            dcs = DistributionCenter.objects.select_related('provincial_office').all().order_by('provincial_office__name', 'name')
        else:
            po = user.provincial_office
            dcs = DistributionCenter.objects.filter(provincial_office=po).order_by('name') if po else DistributionCenter.objects.none()

        MONTHS = [
            (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
            (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
            (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
        ]

        # Build current targets map: {(dc_id, month): target_loss_percent}
        existing = {}
        if active_fy:
            for t in DCMonthlyTarget.objects.filter(fiscal_year=active_fy, distribution_center__in=dcs):
                existing[(t.distribution_center_id, t.month)] = float(t.target_loss_percent)

        dc_rows = []
        for dc in dcs:
            month_targets = []
            for m_num, m_name in MONTHS:
                month_targets.append({
                    'month': m_num,
                    'month_name': m_name,
                    'target': existing.get((dc.pk, m_num), ''),
                })
            dc_rows.append({'dc': dc, 'month_targets': month_targets})

        return render(request, self.template_name, {
            'active_fy': active_fy,
            'dc_rows': dc_rows,
            'months': MONTHS,
        })

    def post(self, request):
        user = request.user
        active_fy = FiscalYear.objects.filter(is_active=True).first()
        if not active_fy:
            messages.error(request, 'No active fiscal year found.')
            return redirect('dc_monthly_targets')

        if getattr(user, 'is_system_admin', False) or user.is_top_management:
            allowed_dcs = set(DistributionCenter.objects.values_list('pk', flat=True))
        else:
            po = user.provincial_office
            allowed_dcs = set(DistributionCenter.objects.filter(provincial_office=po).values_list('pk', flat=True)) if po else set()

        saved = 0
        for key, val in request.POST.items():
            # key format: target_<dc_id>_<month>
            if not key.startswith('target_'):
                continue
            parts = key.split('_')
            if len(parts) != 3:
                continue
            try:
                dc_id = int(parts[1])
                month = int(parts[2])
                val = val.strip()
                if not val:
                    # Delete existing target if blank submitted
                    DCMonthlyTarget.objects.filter(
                        distribution_center_id=dc_id, fiscal_year=active_fy, month=month
                    ).delete()
                    continue
                target_pct = float(val)
            except (ValueError, TypeError):
                continue

            if dc_id not in allowed_dcs:
                continue
            if not (1 <= month <= 12):
                continue

            DCMonthlyTarget.objects.update_or_create(
                distribution_center_id=dc_id,
                fiscal_year=active_fy,
                month=month,
                defaults={'target_loss_percent': target_pct, 'set_by': user},
            )
            saved += 1

        messages.success(request, f'Saved {saved} monthly target(s) successfully.')
        return redirect('dc_monthly_targets')


# ─────────────────────────── USER MANAGEMENT ───────────────────────────

class UserListView(LoginRequiredMixin, View):
    template_name = 'nea_loss/users/list.html'

    def get(self, request):
        if not request.user.is_system_admin and not request.user.is_top_management and not request.user.is_staff:
            return redirect('dashboard')
        users = NEAUser.objects.select_related('provincial_office', 'distribution_center').all()
        return render(request, self.template_name, {'users': users})


class UserCreateView(LoginRequiredMixin, View):
    template_name = 'nea_loss/users/create.html'

    def get(self, request):
        if not request.user.is_system_admin and not request.user.is_top_management and not request.user.is_staff:
            return redirect('dashboard')
        return render(request, self.template_name, {
            'role_choices': NEAUser.ROLE_CHOICES,
            'provincial_offices': ProvincialOffice.objects.all(),
            'distribution_centers': DistributionCenter.objects.all(),
        })

    def post(self, request):
        username = request.POST.get('username')
        email = request.POST.get('email')
        full_name = request.POST.get('full_name')
        role = request.POST.get('role')
        password = request.POST.get('password')
        po_id = request.POST.get('provincial_office') or None
        dc_id = request.POST.get('distribution_center') or None

        if NEAUser.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return redirect('user_create')

        user = NEAUser.objects.create_user(
            username=username, email=email, full_name=full_name,
            role=role, password=password,
            provincial_office_id=po_id, distribution_center_id=dc_id,
            employee_id=request.POST.get('employee_id', ''),
            phone=request.POST.get('phone', ''),
            designation=request.POST.get('designation', ''),
        )
        messages.success(request, f'User {username} created successfully.')
        return redirect('user_list')


class UserEditView(LoginRequiredMixin, View):
    template_name = 'nea_loss/users/edit.html'

    def get(self, request, pk):
        if not request.user.is_system_admin and not request.user.is_top_management and not request.user.is_staff:
            return redirect('dashboard')
        user = get_object_or_404(NEAUser, pk=pk)
        return render(request, self.template_name, {
            'edit_user': user,
            'role_choices': NEAUser.ROLE_CHOICES,
            'provincial_offices': ProvincialOffice.objects.all(),
            'distribution_centers': DistributionCenter.objects.all(),
        })

    def post(self, request, pk):
        user = get_object_or_404(NEAUser, pk=pk)
        user.full_name = request.POST.get('full_name', user.full_name)
        user.role = request.POST.get('role', user.role)
        user.email = request.POST.get('email', user.email)
        user.phone = request.POST.get('phone', user.phone)
        user.designation = request.POST.get('designation', user.designation)
        po_id = request.POST.get('provincial_office') or None
        dc_id = request.POST.get('distribution_center') or None
        user.provincial_office_id = po_id
        user.distribution_center_id = dc_id
        user.is_active = request.POST.get('is_active') == 'on'
        user.save()
        messages.success(request, 'User updated successfully.')
        return redirect('user_list')


# ─────────────────────────── API VIEWS ───────────────────────────

@login_required
def api_dashboard_chart(request):
    active_fy = FiscalYear.objects.filter(is_active=True).first()
    if not active_fy:
        return JsonResponse({'months': [], 'received': [], 'utilised': [], 'loss': []})

    user = request.user
    # Only APPROVED reports feed the dashboard chart for management views
    if user.is_dc_level and user.distribution_center:
        reports = LossReport.objects.filter(
            fiscal_year=active_fy,
            distribution_center=user.distribution_center,
            status__in=['SUBMITTED', 'PROVINCIAL_REVIEWED', 'APPROVED'],
        )
    elif user.is_provincial and user.provincial_office:
        reports = LossReport.objects.filter(
            fiscal_year=active_fy,
            status='APPROVED',
            distribution_center__provincial_office=user.provincial_office,
        )
    else:
        reports = LossReport.objects.filter(fiscal_year=active_fy, status='APPROVED')

    data = {}
    for report in reports:
        for md in report.monthly_data.order_by('month'):
            m = md.month_name
            if m not in data:
                data[m] = {'received': 0, 'utilised': 0, 'loss': 0}
            data[m]['received'] += float(md.net_energy_received)
            data[m]['utilised'] += float(md.total_energy_utilised)
            data[m]['loss'] += float(md.loss_unit)

    if not data:
        return JsonResponse({'months': [], 'received': [], 'utilised': [], 'loss': []})

    months = list(data.keys())
    return JsonResponse({
        'months': months,
        'received': [data[m]['received'] for m in months],
        'utilised': [data[m]['utilised'] for m in months],
        'loss': [data[m]['loss'] for m in months],
    })


@login_required
def api_loss_summary(request):
    active_fy = FiscalYear.objects.filter(is_active=True).first()
    reports = LossReport.objects.filter(fiscal_year=active_fy) if active_fy else LossReport.objects.none()
    total_received = reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0
    total_loss = reports.aggregate(s=Sum('total_loss_kwh'))['s'] or 0
    loss_pct = float(total_loss / total_received * 100) if total_received > 0 else 0
    return JsonResponse({
        'total_received_mwh': round(float(total_received) / 1000, 2),
        'total_loss_mwh': round(float(total_loss) / 1000, 2),
        'loss_pct': round(loss_pct, 2),
        'target_pct': float(active_fy.loss_target_percent) if active_fy else 3.35,
    })


@login_required
def api_mark_notifications_read(request):
    if request.method == 'POST':
        Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
        return JsonResponse({'success': True})
    return JsonResponse({'error': 'POST required'})


@login_required
def api_create_monthly_data(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'})
    try:
        data = json.loads(request.body)
        report_id = data.get('report_id')
        month = data.get('month')
        
        if not report_id or not month:
            return JsonResponse({'error': 'report_id and month are required'}, status=400)
            
        report = get_object_or_404(LossReport, pk=report_id)
        
        # Check if user can edit this report
        if not _can_edit_report(request.user, report):
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        month_names = {
            1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
            5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
            9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
        }

        # Create monthly data
        monthly, created = MonthlyLossData.objects.get_or_create(
            report=report,
            month=month,
            defaults={
                'month_name': month_names.get(int(month), ''),
            }
        )

        return JsonResponse({
            'success': True,
            'monthly_id': monthly.pk,
            'created': created
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_save_meter_readings(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'})
    try:
        data = json.loads(request.body)
        monthly_id = data.get('monthly_id')
        readings = data.get('readings', [])
        
        if not monthly_id:
            return JsonResponse({'error': 'monthly_id is required'}, status=400)
            
        # Get the monthly data, return error if doesn't exist
        try:
            monthly = MonthlyLossData.objects.get(pk=monthly_id)
        except MonthlyLossData.DoesNotExist:
            return JsonResponse({'error': 'Monthly data not found. Please enter some data first.'}, status=400)
            
        if not _can_edit_report(request.user, monthly.report):
            return JsonResponse({'error': 'Permission denied'}, status=403)

        import_types = {'SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP', 'ENERGY_IMPORT'}
        export_types = {'EXPORT_DC', 'EXPORT_IPP', 'ENERGY_EXPORT'}

        for r in readings:
            mp = get_object_or_404(MeterPoint, pk=r['meter_point_id'])
            if mp.distribution_center_id != monthly.report.distribution_center_id:
                return JsonResponse({'error': 'Invalid meter point'}, status=400)

            if mp.is_single_reading:
                # ENERGY_IMPORT / ENERGY_EXPORT: only present_reading matters; previous=0
                mr, created = MeterReading.objects.update_or_create(
                    monthly_data=monthly, meter_point=mp,
                    defaults={
                        'present_reading': decimal.Decimal(str(r['present_reading'])),
                        'previous_reading': decimal.Decimal('0'),
                        'multiplying_factor': decimal.Decimal(str(r.get('multiplying_factor', mp.multiplying_factor))),
                    }
                )
                if not created:
                    # update_or_create uses queryset.update() on existing rows,
                    # bypassing the custom save() that recalculates unit_kwh.
                    # Call save() explicitly so unit_kwh is always recalculated.
                    mr.save()
            else:
                # For regular feeders: if previous_reading is 0 / not provided and this is
                # not Shrawan, auto-fill from the last approved month's present reading.
                provided_prev = decimal.Decimal(str(r.get('previous_reading', 0) or 0))
                report_month = monthly.report.month

                if provided_prev == 0 and report_month > 1:
                    # Look up last month's approved present reading for this meter point
                    prev_month_num = report_month - 1
                    auto_prev = MeterReading.objects.filter(
                        meter_point=mp,
                        monthly_data__report__distribution_center=monthly.report.distribution_center,
                        monthly_data__report__fiscal_year=monthly.report.fiscal_year,
                        monthly_data__report__status='APPROVED',
                        monthly_data__month=prev_month_num,
                    ).values_list('present_reading', flat=True).first()
                    if auto_prev is not None:
                        provided_prev = auto_prev

                mr, created = MeterReading.objects.update_or_create(
                    monthly_data=monthly, meter_point=mp,
                    defaults={
                        'present_reading': decimal.Decimal(str(r['present_reading'])),
                        'previous_reading': provided_prev,
                        'multiplying_factor': decimal.Decimal(str(r.get('multiplying_factor', mp.multiplying_factor))),
                    }
                )
                if not created:
                    # update_or_create uses queryset.update() on existing rows,
                    # bypassing the custom save() that recalculates unit_kwh.
                    # Call save() explicitly so unit_kwh is always recalculated.
                    mr.save()
        total_import = decimal.Decimal('0')
        total_export = decimal.Decimal('0')
        for mr in monthly.meter_readings.select_related('meter_point').all():
            if mr.meter_point.source_type in import_types:
                total_import += mr.unit_kwh
            elif mr.meter_point.source_type in export_types:
                total_export += mr.unit_kwh

        monthly.total_energy_import = total_import
        monthly.total_energy_export = total_export
        monthly.net_energy_received = total_import - total_export
        monthly.loss_unit = monthly.net_energy_received - monthly.total_energy_utilised
        if monthly.net_energy_received > 0:
            monthly.monthly_loss_percent = monthly.loss_unit / monthly.net_energy_received
        else:
            monthly.monthly_loss_percent = decimal.Decimal('0')
        monthly.save()
        monthly.report.calculate_summary()

        return JsonResponse({
            'success': True,
            'total_import': float(total_import),
            'total_export': float(total_export),
            'net_received': float(monthly.net_energy_received),
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_save_consumer_data(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'})
    try:
        data = json.loads(request.body)
        monthly_id = data.get('monthly_id')
        utilisations = data.get('utilisations', [])
        counts = data.get('counts', [])
        
        if not monthly_id:
            return JsonResponse({'error': 'monthly_id is required'}, status=400)
            
        # Get the monthly data, return error if doesn't exist
        try:
            monthly = MonthlyLossData.objects.get(pk=monthly_id)
        except MonthlyLossData.DoesNotExist:
            return JsonResponse({'error': 'Monthly data not found. Please enter some data first.'}, status=400)
            
        if not _can_edit_report(request.user, monthly.report):
            return JsonResponse({'error': 'Permission denied'}, status=403)
        dc_id = monthly.report.distribution_center_id

        for u in utilisations:
            cat = get_object_or_404(ConsumerCategory, pk=u['category_id'])
            if cat.distribution_center_id not in (None, dc_id):
                return JsonResponse({'error': 'Invalid consumer category'}, status=400)
            eu, _ = EnergyUtilisation.objects.update_or_create(
                monthly_data=monthly, consumer_category=cat,
                defaults={
                    'energy_kwh': decimal.Decimal(str(u['energy_kwh'])),
                    'remarks': (u.get('remarks') or '')[:200],
                }
            )

        for c in counts:
            cat = get_object_or_404(ConsumerCategory, pk=c['category_id'])
            if cat.distribution_center_id not in (None, dc_id):
                return JsonResponse({'error': 'Invalid consumer category'}, status=400)
            ConsumerCount.objects.update_or_create(
                monthly_data=monthly, consumer_category=cat,
                defaults={
                    'count': int(c['count']),
                    'remarks': (c.get('remarks') or '')[:200],
                }
            )

        total_utilised = monthly.energy_utilisations.aggregate(s=Sum('energy_kwh'))['s'] or decimal.Decimal('0')
        monthly.total_energy_utilised = total_utilised
        monthly.loss_unit = monthly.net_energy_received - total_utilised
        if monthly.net_energy_received > 0:
            monthly.monthly_loss_percent = monthly.loss_unit / monthly.net_energy_received
        monthly.save()
        monthly.report.calculate_summary()

        return JsonResponse({
            'success': True,
            'total_utilised': float(total_utilised),
            'loss_unit': float(monthly.loss_unit),
            'loss_pct': float(monthly.monthly_loss_percent) * 100,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_POST
def api_manage_meter_point(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    report = get_object_or_404(LossReport, pk=data.get('report_pk'))
    if not _can_edit_report(request.user, report):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    action = data.get('action')
    if action == 'create':
        name = (data.get('name') or '').strip()
        if not name:
            return JsonResponse({'error': 'Name is required'}, status=400)
        source_type = data.get('source_type')
        valid_types = {k for k, _ in MeterPoint.SOURCE_TYPE_CHOICES}
        if source_type not in valid_types:
            return JsonResponse({'error': 'Invalid source type'}, status=400)
        
        # Define month names for this function
        month_names = {
            1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
            5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
            9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
        }
        
        mp = MeterPoint.objects.create(
            distribution_center=report.distribution_center,
            name=name,
            code='',
            source_type=source_type,
            voltage_level=(data.get('voltage_level') or '').strip()[:20],
            multiplying_factor=decimal.Decimal(str(data.get('multiplying_factor') or 1)),
            is_active=True,
        )
        
        # Automatically mark this meter point as inactive for previous months
        # to prevent it from appearing in historical reports
        current_month = report.month  # Current month being viewed
        for month_num in range(1, current_month):
            # Find or create monthly data for this month
            monthly_data, _ = MonthlyLossData.objects.get_or_create(
                report=report,
                month=month_num,
                defaults={'month_name': month_names.get(month_num, '')}
            )
            # Mark as inactive for previous months
            MonthlyMeterPointStatus.objects.update_or_create(
                monthly_data=monthly_data,
                meter_point=mp,
                defaults={'is_active': False}
            )
        
        return JsonResponse({
            'success': True,
            'meter_point': {
                'id': mp.pk,
                'name': mp.name,
                'source_type': mp.source_type,
                'source_type_display': mp.get_source_type_display(),
                'voltage_level': mp.voltage_level,
                'multiplying_factor': float(mp.multiplying_factor),
            },
        })

    if action == 'delete':
        mp = get_object_or_404(MeterPoint, pk=data.get('meter_point_id'))
        if mp.distribution_center_id != report.distribution_center_id:
            return JsonResponse({'error': 'Invalid meter point'}, status=400)

        monthly_id = data.get('monthly_id')
        if not monthly_id:
            return JsonResponse({'error': 'monthly_id is required for delete'}, status=400)

        monthly = get_object_or_404(MonthlyLossData, pk=monthly_id)

        # ALWAYS soft-delete: mark this feeder inactive for THIS month only.
        # The MeterPoint record is NEVER deleted — this preserves all data in
        # every previously approved report that references this feeder.
        # Only the current month's reading is removed.
        MonthlyMeterPointStatus.objects.update_or_create(
            monthly_data=monthly,
            meter_point=mp,
            defaults={'is_active': False},
        )
        # Remove the reading for this month only (approved months are untouched)
        MeterReading.objects.filter(monthly_data=monthly, meter_point=mp).delete()

        message = (
            f'"{mp.name}" removed from {monthly.month_name} only. '
            f'All other months (including approved reports) are unaffected.'
        )

        # Recalculate totals for current month
        import_types = {'SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP', 'ENERGY_IMPORT'}
        export_types = {'EXPORT_DC', 'EXPORT_IPP', 'ENERGY_EXPORT'}
        total_import = decimal.Decimal('0')
        total_export = decimal.Decimal('0')
        for mr in monthly.meter_readings.select_related('meter_point').all():
            if mr.meter_point.source_type in import_types:
                total_import += mr.unit_kwh
            elif mr.meter_point.source_type in export_types:
                total_export += mr.unit_kwh
        monthly.total_energy_import = total_import
        monthly.total_energy_export = total_export
        monthly.net_energy_received = total_import - total_export
        monthly.loss_unit = monthly.net_energy_received - monthly.total_energy_utilised
        if monthly.net_energy_received > 0:
            monthly.monthly_loss_percent = monthly.loss_unit / monthly.net_energy_received
        else:
            monthly.monthly_loss_percent = decimal.Decimal('0')
        monthly.save()
        monthly.report.calculate_summary()

        return JsonResponse({
            'success': True,
            'action_taken': 'disabled_for_month',
            'message': message,
        })

    return JsonResponse({'error': 'Unknown action'}, status=400)


@login_required
@require_POST
def api_manage_consumer_category(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    report = get_object_or_404(LossReport, pk=data.get('report_pk'))
    if not _can_edit_report(request.user, report):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    action = data.get('action')
    if action == 'create':
        name = (data.get('name') or '').strip()
        if not name:
            return JsonResponse({'error': 'Name is required'}, status=400)
        dc = report.distribution_center
        for _ in range(5):
            code = f'DC{dc.pk}_{uuid.uuid4().hex[:12]}'.upper()
            code = code[:40]
            if not ConsumerCategory.objects.filter(code=code).exists():
                break
        else:
            return JsonResponse({'error': 'Could not allocate category code'}, status=500)
        cat = ConsumerCategory.objects.create(
            name=name,
            code=code,
            distribution_center=dc,
            display_order=500,
        )
        return JsonResponse({
            'success': True,
            'category': {'id': cat.pk, 'name': cat.name},
        })

    if action == 'delete':
        cat = get_object_or_404(ConsumerCategory, pk=data.get('category_id'))
        if cat.distribution_center_id != report.distribution_center_id:
            return JsonResponse({
                'error': 'Only categories added by your DC can be removed here.',
            }, status=400)
        cat.delete()
        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Unknown action'}, status=400)


@login_required
def api_recalculate(request, report_pk):
    report = get_object_or_404(LossReport, pk=report_pk)
    if not _can_edit_report(request.user, report):
        return JsonResponse({"error": "Permission denied"}, status=403)

    # Recalculate this report
    report.calculate_summary()

    # CASCADE: if this is an APPROVED report, recalculate all subsequent months
    # in the same DC + fiscal year so cumulative figures stay consistent.
    if report.status == "APPROVED":
        subsequent_reports = LossReport.objects.filter(
            distribution_center=report.distribution_center,
            fiscal_year=report.fiscal_year,
            month__gt=report.month,
            status="APPROVED",
        ).order_by("month")
        for subsequent in subsequent_reports:
            subsequent.calculate_summary()

    return JsonResponse({
        "success": True,
        "cumulative_loss_pct": float(report.cumulative_loss_percent) * 100,
        "total_received": float(report.total_received_kwh),
        "total_loss": float(report.total_loss_kwh),
    })


# ─────────────────────────── HELPER FUNCTIONS ───────────────────────────

def _can_create_loss_report(user):
    """Only DC staff/managers create DC loss reports.
    MD/DMD/Director/Provincial are view-only for DC reports.
    System Admin can do everything."""
    if not user.is_authenticated:
        return False
    if getattr(user, 'is_system_admin', False):
        return True
    if user.is_dc_level:
        return bool(getattr(user, 'distribution_center_id', None))
    # Provincial, MD, DMD, Director cannot create DC-level loss reports
    return False


def _can_edit_report(user, report):
    if getattr(user, 'is_system_admin', False):
        return True
    # Top management (MD/DMD/Director) can edit APPROVED reports to make corrections
    if user.is_top_management:
        return report.status == 'APPROVED'
    if user.is_provincial:
        return False
    if report.status not in ['DRAFT', 'REJECTED']:
        return False
    if user.is_dc_level:
        dc = getattr(user, 'distribution_center', None)
        if dc and dc.pk == report.distribution_center_id:
            return True
    return False


def _can_view_report(user, report):
    if getattr(user, 'is_system_admin', False):
        return True
    if user.is_top_management:
        return True
    if user.is_provincial:
        po = getattr(user, 'provincial_office', None)
        if po and po.pk == report.distribution_center.provincial_office_id:
            return True
    if user.is_dc_level:
        dc = getattr(user, 'distribution_center', None)
        if dc and dc.pk == report.distribution_center_id:
            return True
    return False


def _can_approve_report(user):
    return getattr(user, 'is_system_admin', False) or user.is_top_management or user.is_provincial



# ─────────────────────────── MESSAGING VIEWS ───────────────────────────

# Messaging
class MessageInboxView(LoginRequiredMixin, View):
    template_name = 'nea_loss/users/inbox.html'

    def get(self, request):
        inbox = Message.objects.filter(recipient=request.user).select_related('sender').order_by('-created_at')
        sent = Message.objects.filter(sender=request.user).select_related('recipient').order_by('-created_at')
        unread_count = inbox.filter(is_read=False).count()
        
        return render(request, self.template_name, {
            'inbox': inbox,
            'sent': sent,
            'unread_count': unread_count,
        })


class MessageComposeView(LoginRequiredMixin, View):
    template_name = 'nea_loss/users/message_compose.html'

    def get(self, request):
        # Get all users except current user
        users = NEAUser.objects.exclude(pk=request.user.pk).order_by('full_name')
        return render(request, self.template_name, {
            'users': users,
        })

    def post(self, request):
        recipient_id = request.POST.get('recipient')
        subject = request.POST.get('subject')
        body = request.POST.get('body')

        if not recipient_id or not subject or not body:
            messages.error(request, 'Please fill in all fields.')
            return self.get(request)

        try:
            recipient = NEAUser.objects.get(pk=recipient_id)
            message = Message.objects.create(
                sender=request.user,
                recipient=recipient,
                subject=subject,
                body=body
            )
            messages.success(request, 'Message sent successfully.')
            return redirect('message_inbox')
        except NEAUser.DoesNotExist:
            messages.error(request, 'Recipient not found.')
            return self.get(request)


class MessageDetailView(LoginRequiredMixin, View):
    template_name = 'nea_loss/reports/message_detail.html'

    def get(self, request, pk):
        message = get_object_or_404(Message, Q(pk=pk) & (Q(sender=request.user) | Q(recipient=request.user)))
        
        # Mark as read if recipient
        if message.recipient == request.user and not message.is_read:
            message.is_read = True
            message.save()
        
        # Get replies
        replies = Message.objects.filter(parent=message).order_by('created_at')
        
        return render(request, self.template_name, {
            'msg': message,
            'replies': replies,
        })


@login_required
def message_delete(request, pk):
    message = get_object_or_404(Message, Q(pk=pk) & (Q(sender=request.user) | Q(recipient=request.user)))
    message.delete()
    messages.success(request, 'Message deleted successfully.')
    return redirect('message_inbox')


@login_required
def message_reply(request, pk):
    parent_message = get_object_or_404(Message, Q(pk=pk) & (Q(sender=request.user) | Q(recipient=request.user)))
    
    if request.method == 'POST':
        body = request.POST.get('body')
        if not body:
            messages.error(request, 'Please enter a message.')
            return redirect('message_detail', pk=pk)
        
        # Determine recipient (reply to the other person)
        recipient = parent_message.sender if parent_message.recipient == request.user else parent_message.recipient
        
        reply = Message.objects.create(
            sender=request.user,
            recipient=recipient,
            subject=f"Re: {parent_message.subject}",
            body=body,
            parent=parent_message
        )
        
        messages.success(request, 'Reply sent successfully.')
        return redirect('message_detail', pk=pk)


@login_required
def api_unread_messages(request):
    count = Message.objects.filter(recipient=request.user, is_read=False).count()
    return JsonResponse({'unread': count})


# ─────────────────────────── PROVINCIAL REPORT VIEWS ───────────────────────────

class ProvincialReportCreateView(LoginRequiredMixin, View):
    """Provincial office generates monthly consolidated report from DC data."""
    template_name = 'nea_loss/reports/provincial_create.html'

    def dispatch(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            if not (getattr(user, 'is_system_admin', False) or user.is_provincial):
                messages.error(request, 'Only Provincial Office users can create provincial reports.')
                return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        user = request.user
        if getattr(user, 'is_system_admin', False):
            provincial_offices = ProvincialOffice.objects.all()
        else:
            provincial_offices = ProvincialOffice.objects.filter(pk=user.provincial_office_id)

        return render(request, self.template_name, {
            'fiscal_years': FiscalYear.objects.all(),
            'provincial_offices': provincial_offices,
            'months_list': [
                (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
                (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
                (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
            ],
        })

    def post(self, request):
        fy_id = request.POST.get('fiscal_year')
        month_str = request.POST.get('month', '')
        # Handle "Report Till Now" case (month=0) as show_all=True
        if month_str.strip() == '0':
            month = 0
        else:
            month = int(month_str) if month_str.strip() else None
        po_id = request.POST.get('provincial_office')
        action = request.POST.get('action', 'preview')

        try:
            fy = FiscalYear.objects.get(pk=fy_id)
            po = ProvincialOffice.objects.get(pk=po_id)
        except Exception:
            messages.error(request, 'Invalid fiscal year or provincial office.')
            return redirect('provincial_report_create')

        month_names = {
            1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
            5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
            9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
        }

        # Auto-determine if showing all months or specific month
        show_all = month is None or month == 0  # If no month selected, show all available months
        
        # If no month selected, find the latest available month
        if show_all:
            latest_report = LossReport.objects.filter(
                distribution_center__provincial_office=po,
                fiscal_year=fy,
                status='APPROVED'
            ).order_by('-month').first()
            
            if latest_report:
                month = latest_report.month
            else:
                messages.error(request, 'No approved reports found for this provincial office and fiscal year.')
                return redirect('provincial_report_create')

        # show_all=True -> cumulative Shrawan->month view; False -> selected month only

        # Gather all DC reports under this provincial office for this FY/month
        dcs = DistributionCenter.objects.filter(provincial_office=po)
        dc_report_data = []
        grand_total_received = 0
        grand_total_utilised = 0

        for dc in dcs:
            # ── Month filter logic ──
            # "all_months" mode: include Shrawan up to selected month (cumulative view)
            # normal mode: show ONLY the selected month
            if show_all:
                dc_reports_range = LossReport.objects.filter(
                    distribution_center=dc,
                    fiscal_year=fy,
                    month__lte=month
                ).order_by('month')
            else:
                dc_reports_range = LossReport.objects.filter(
                    distribution_center=dc,
                    fiscal_year=fy,
                    month=month
                ).order_by('month')

            # Month-specific report (always the selected month)
            month_report = LossReport.objects.filter(
                distribution_center=dc, fiscal_year=fy, month=month
            ).first()

            # For cumulative calculation always use Shrawan → selected month
            dc_reports_ytd = LossReport.objects.filter(
                distribution_center=dc,
                fiscal_year=fy,
                month__lte=month
            ).order_by('month')

            month_received = float(month_report.total_received_kwh) if month_report else 0
            month_utilised = float(month_report.total_utilised_kwh) if month_report else 0

            # Cumulative = Σloss_so_far / Σreceived_so_far × 100 (same formula as dashboard)
            ytd_received = sum(float(r.total_received_kwh) for r in dc_reports_ytd)
            ytd_utilised = sum(float(r.total_utilised_kwh) for r in dc_reports_ytd)
            ytd_loss = ytd_received - ytd_utilised
            ytd_cl = round(ytd_loss / ytd_received * 100, 4) if ytd_received else 0

            # Monthly loss % = loss of that month only / received of that month × 100
            monthly_il = round((month_received - month_utilised) / month_received * 100, 4) if month_received else 0

            # Monthly breakdown: only for displayed range
            monthly_breakdown = {}
            for r in dc_reports_range:
                mn = r.month
                md_loss = float(r.total_received_kwh) - float(r.total_utilised_kwh)
                md_il = round(md_loss / float(r.total_received_kwh) * 100, 4) if float(r.total_received_kwh) else 0
                monthly_breakdown[mn] = {
                    'received': float(r.total_received_kwh),
                    'utilised': float(r.total_utilised_kwh),
                    'monthly_il': md_il,
                }

            # For show_all mode, use YTD cumulative totals; for specific month, use month values
            if show_all:
                grand_total_received += ytd_received
                grand_total_utilised += ytd_utilised
            else:
                grand_total_received += month_received
                grand_total_utilised += month_utilised

            # ── DC-specific provincial yearly target ──
            dc_prov_target = DCYearlyTarget.objects.filter(
                distribution_center=dc, fiscal_year=fy
            ).first()
            dc_target_pct = float(dc_prov_target.target_loss_percent) if dc_prov_target else None

            dc_report_data.append({
                'dc': dc,
                'month_received': month_received,
                'month_utilised': month_utilised,
                'monthly_il': monthly_il,
                'ytd_received': ytd_received,
                'ytd_utilised': ytd_utilised,
                'ytd_cl': ytd_cl,
                'monthly_breakdown': monthly_breakdown,
                'report_status': month_report.status if month_report else 'NO_REPORT',
                'dc_target': dc_target_pct,    # Provincial target for this DC this month
                'nea_target': float(fy.loss_target_percent),  # NEA target (for reference only)
            })

        grand_loss = grand_total_received - grand_total_utilised
        grand_il = round(grand_loss / grand_total_received * 100, 4) if grand_total_received else 0

        if action == 'save':
            # Save/update the provincial report record
            prov_report, created = ProvincialReport.objects.get_or_create(
                provincial_office=po,
                fiscal_year=fy,
                month=month,
                defaults={'created_by': request.user, 'status': 'DRAFT'}
            )
            AuditLog.objects.create(
                user=request.user,
                action='CREATE' if created else 'UPDATE',
                model_name='ProvincialReport',
                object_id=prov_report.pk,
                description=f"{'Created' if created else 'Updated'} provincial report for {po.name} - {fy.year_bs} - {month_names.get(month,'')}",
            )
            messages.success(request, f'Provincial report for {month_names.get(month,"")} saved successfully.')

        user = request.user
        if getattr(user, 'is_system_admin', False):
            provincial_offices = ProvincialOffice.objects.all()
        else:
            provincial_offices = ProvincialOffice.objects.filter(pk=user.provincial_office_id)

        context = {
            'fiscal_years': FiscalYear.objects.all(),
            'provincial_offices': provincial_offices,
            'months_list': [
                (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
                (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
                (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
            ],
            'report_data': dc_report_data,
            'grand_total_received': grand_total_received,
            'grand_total_utilised': grand_total_utilised,
            'grand_il': round(grand_il, 2),
            'selected_fy': fy,
            'selected_month': month,
            'selected_month_name': month_names.get(month, ''),
            'selected_po': po,
            'nea_target_pct': float(fy.loss_target_percent),
            'show_all': show_all,
            'months_range': list(range(1, month + 1)) if show_all else [month],
            'month_names': month_names,
        }
        return render(request, self.template_name, context)



class ProvincialDCReportsView(LoginRequiredMixin, View):
    """For Provincial users: shows all DCs in their province with month selector.
    Shows which DCs have submitted a report for the selected month."""
    template_name = 'nea_loss/reports/provincial_dc_reports.html'

    def get(self, request):
        user = request.user
        if not (user.is_provincial and user.provincial_office) and not getattr(user, 'is_system_admin', False):
            return redirect('dashboard')

        active_fy = FiscalYear.objects.filter(is_active=True).first()
        fiscal_years = FiscalYear.objects.all().order_by('-year_ad_start')

        # Filters
        selected_fy_id = request.GET.get('fiscal_year')
        selected_month = request.GET.get('month', '')
        try:
            selected_month = int(selected_month)
        except (ValueError, TypeError):
            selected_month = ''

        selected_fy = active_fy
        if selected_fy_id:
            try:
                selected_fy = FiscalYear.objects.get(pk=selected_fy_id)
            except FiscalYear.DoesNotExist:
                pass

        if user.is_provincial and user.provincial_office:
            dcs = DistributionCenter.objects.filter(
                provincial_office=user.provincial_office
            ).order_by('name')
        else:
            dcs = DistributionCenter.objects.all().order_by('name')

        MONTH_CHOICES = [
            (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
            (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
            (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
        ]

        dc_rows = []
        for dc in dcs:
            row = {'dc': dc, 'report': None, 'status': 'Not Submitted'}
            if selected_fy and selected_month:
                report = LossReport.objects.filter(
                    distribution_center=dc,
                    fiscal_year=selected_fy,
                    month=selected_month,
                ).first()
                if report:
                    row['report'] = report
                    row['status'] = report.get_status_display()
                else:
                    row['status'] = 'Not Submitted'
            dc_rows.append(row)

        submitted_count = sum(1 for r in dc_rows if r['report'])
        not_submitted_count = len(dc_rows) - submitted_count

        return render(request, self.template_name, {
            'dc_rows': dc_rows,
            'fiscal_years': fiscal_years,
            'selected_fy': selected_fy,
            'selected_month': selected_month,
            'month_choices': MONTH_CHOICES,
            'provincial_office': getattr(user, 'provincial_office', None),
            'submitted_count': submitted_count,
            'not_submitted_count': not_submitted_count,
        })

class ProvincialReportPrintView(LoginRequiredMixin, View):
    """Print provincial monthly report in formal format."""
    template_name = 'nea_loss/reports/provincial_print.html'

    def dispatch(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            if not (getattr(user, 'is_system_admin', False) or user.is_provincial):
                messages.error(request, 'Only provincial managers can print provincial reports.')
                return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        user = request.user
        if getattr(user, 'is_system_admin', False):
            provincial_offices = ProvincialOffice.objects.all()
        else:
            provincial_offices = ProvincialOffice.objects.filter(pk=user.provincial_office_id)

        return render(request, self.template_name, {
            'provincial_offices': provincial_offices,
            'fiscal_years': FiscalYear.objects.all(),
            'months_list': [
                (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
                (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
                (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
            ],
        })

    def post(self, request):
        fy_id = request.POST.get('fiscal_year')
        month_str = request.POST.get('month', '')
        po_id = request.POST.get('provincial_office')
        
        try:
            fy = FiscalYear.objects.get(pk=fy_id)
            po = ProvincialOffice.objects.get(pk=po_id)
            # Handle "Report Till Now" case (month=0) as show_all=True
            if month_str.strip() == '0':
                month = 0
            else:
                month = int(month_str) if month_str.strip() else None
        except Exception:
            messages.error(request, 'Invalid fiscal year or provincial office.')
            return redirect('provincial_report_create')

        # Auto-determine if showing all months or specific month
        show_all = month is None or month == 0  # If no month selected, show all available months
        
        # If no month selected, find latest available month
        if show_all:
            latest_report = LossReport.objects.filter(
                distribution_center__provincial_office=po,
                fiscal_year=fy,
                status='APPROVED'
            ).order_by('-month').first()
            
            if latest_report:
                month = latest_report.month
            else:
                messages.error(request, 'No approved reports found for this provincial office and fiscal year.')
                return redirect('provincial_report_create')

        # Get all distribution centers under this provincial office
        distribution_centers = DistributionCenter.objects.filter(provincial_office=po)
        
        # Prepare monthly data for each DC
        months_range = list(range(1, month + 1)) if show_all else [month]
        months_list = [
            (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
            (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
            (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
        ]
        
        # Build monthly data structure for each DC
        dc_monthly_data = []
        for dc in distribution_centers:
            dc_data = []
            for m in months_range:
                month_name = dict(months_list).get(m, '')
                
                # Get meter points data
                import_meter_points = dc.meter_points.filter(
                    source_type__in=["SUBSTATION","FEEDER_11KV","FEEDER_33KV","INTERBRANCH","IPP","ENERGY_IMPORT"]
                )
                export_meter_points = dc.meter_points.filter(
                    source_type__in=["EXPORT_DC","EXPORT_IPP","ENERGY_EXPORT"]
                )
                consumer_categories = dc.consumer_categories.all()
                
                # Calculate totals
                total_energy_import = 0
                total_energy_export = 0
                total_energy_utilised = 0
                total_loss = 0
                
                # Get monthly report if exists
                monthly_report = LossReport.objects.filter(
                    distribution_center=dc, fiscal_year=fy, month=m
                ).first()
                
                if monthly_report:
                    total_energy_import = float(monthly_report.total_energy_import or 0)
                    total_energy_export = float(monthly_report.total_energy_export or 0)
                    total_energy_utilised = float(monthly_report.total_utilised_kwh or 0)
                    total_loss = float(monthly_report.total_loss_kwh or 0)
                
                dc_data.append({
                    'month_num': m,
                    'month_name': month_name,
                    'total_energy_import': total_energy_import,
                    'total_energy_export': total_energy_export,
                    'net_energy_received': total_energy_import - total_energy_export,
                    'total_energy_utilised': total_energy_utilised,
                    'loss_unit': total_loss,
                    'monthly_loss_percent': round((total_loss / (total_energy_import - total_energy_export)) * 100, 4) if (total_energy_import - total_energy_export) > 0 else 0,
                    'cumulative_loss_percent': 0,  # Would need cumulative calculation
                    'dc_count': 1,
                    'submitted_count': 1 if monthly_report else 0,
                    'not_submitted_count': 0 if monthly_report else 1,
                })
            
            dc_monthly_data.append({
                'dc': dc,
                'months': dc_data,
                'total_received_kwh': sum(data['net_energy_received'] for data in dc_data),
                'total_utilised_kwh': sum(data['total_energy_utilised'] for Data in dc_data),
                'total_loss_kwh': sum(data['loss_unit'] for Data in dc_data),
                'cumulative_loss_percent': 0,  # Would need proper calculation
                'overall_loss_percent': 0,  # Would need proper calculation
            })
        
        # Calculate provincial totals
        total_dcs = len(distribution_centers)
        total_submitted = sum(1 for dc_data in dc_monthly_data for month_data in dc_data['months'] if month_data.get('submitted_count', 0))
        total_not_submitted = total_dcs - total_submitted

        return render(request, self.template_name, {
            'provincial_office': po,
            'selected_fy': fy,
            'selected_month': month,
            'selected_month_name': dict(months_list).get(month, ''),
            'distribution_centers': dc_monthly_data,
            'months': months_range,
            'total_dcs': total_dcs,
            'total_submitted': total_submitted,
            'total_not_submitted': total_not_submitted,
        })

class ProvincialReportListView(LoginRequiredMixin, View):
    """List of saved provincial reports."""
    template_name = 'nea_loss/reports/provincial_list.html'

    def get(self, request):
        user = request.user
        if getattr(user, 'is_system_admin', False) or user.is_top_management:
            reports = ProvincialReport.objects.select_related('provincial_office', 'fiscal_year', 'created_by').all()
        elif user.is_provincial and user.provincial_office:
            reports = ProvincialReport.objects.filter(provincial_office=user.provincial_office).select_related('provincial_office', 'fiscal_year')
        else:
            reports = ProvincialReport.objects.none()
        return render(request, self.template_name, {
            'reports': reports,
            'fiscal_years': FiscalYear.objects.all(),
        })

class ProvincialReportDetailView(LoginRequiredMixin, View):
    """View saved provincial report details."""
    template_name = 'nea_loss/reports/provincial_detail.html'

    def get(self, request, pk):
        user = request.user
        if not (getattr(user, 'is_system_admin', False) or user.is_provincial):
            messages.error(request, 'Only provincial users can view provincial reports.')
            return redirect('dashboard')
        
        try:
            report = ProvincialReport.objects.select_related(
                'provincial_office', 'fiscal_year', 'created_by'
            ).get(pk=pk)
        except ProvincialReport.DoesNotExist:
            messages.error(request, 'Provincial report not found.')
            return redirect('provincial_report_list')
        
        return render(request, self.template_name, {
            'report': report,
            'fiscal_years': FiscalYear.objects.all(),
        })


def _generate_excel_report(report):
    """Generate the formatted Excel loss report matching NEA format"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loss Analysis Report"

    # Styles
    header_font = Font(name='Arial', bold=True, size=14, color='1B4F72')
    title_font = Font(name='Arial', bold=True, size=11)
    label_font = Font(name='Arial', size=10)
    number_font = Font(name='Arial', size=10)
    blue_fill = PatternFill('solid', start_color='D6EAF8')
    green_fill = PatternFill('solid', start_color='D5F5E3')
    red_fill = PatternFill('solid', start_color='FADBD8')
    gray_fill = PatternFill('solid', start_color='F2F3F4')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center')
    right = Alignment(horizontal='right', vertical='center')

    def thin_border():
        thin = Side(style='thin')
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    # Set column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20
    for col in ['E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 12

    row = 1
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = 'Nepal Electricity Authority'
    ws[f'A{row}'].font = Font(name='Arial', bold=True, size=16, color='1B4F72')
    ws[f'A{row}'].alignment = center

    row += 1
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = f'{report.distribution_center.name}'
    ws[f'A{row}'].font = Font(name='Arial', bold=True, size=13)
    ws[f'A{row}'].alignment = center

    row += 1
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = 'Loss Analysis Report'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = center

    row += 1
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = f'Fiscal Year: {report.fiscal_year.year_bs}'
    ws[f'A{row}'].font = Font(name='Arial', bold=True, size=11)
    ws[f'A{row}'].alignment = center

    row += 1
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = f'Cumulative Loss (%) NEA Target: {report.fiscal_year.loss_target_percent}%'
    # Provincial monthly targets
    row += 1
    ws[f'A{row}'] = 'Provincial Monthly Targets (Loss %)'
    ws[f'A{row}'].font = Font(name='Arial', bold=True, size=10)
    ws[f'A{row}'].fill = PatternFill('solid', start_color='D6EAF8')
    row += 1
    prov_target_headers = ['Month'] + [m.month_name for m in months]
    for ci, h in enumerate(prov_target_headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = Font(name='Arial', bold=True, size=9)
        cell.fill = PatternFill('solid', start_color='EBF5FB')
        cell.alignment = Alignment(horizontal='center')
    row += 1
    prov_targets = DCYearlyTarget.objects.filter(
        distribution_center=report.distribution_center, fiscal_year=report.fiscal_year
    ).first()
    prov_target_pct = float(prov_targets.target_loss_percent) if prov_targets else None
    ws.cell(row=row, column=1, value='Target (%)').font = Font(name='Arial', bold=True, size=9)
    for mi, md in enumerate(months):
        val = prov_target_pct if prov_target_pct else '—'
        cell = ws.cell(row=row, column=2 + mi, value=val)
        cell.font = Font(name='Arial', size=9)
        cell.alignment = Alignment(horizontal='center')
        if isinstance(val, float):
            cell.number_format = '0.000%'
            cell.value = val / 100
    ws[f'A{row}'].font = Font(name='Arial', size=10, color='C0392B')
    ws[f'A{row}'].alignment = center

    row += 2
    months = list(report.monthly_data.order_by('month'))
    month_names = [m.month_name for m in months]

    # Summary header
    headers = ['', '', 'Particular', '', ] + month_names + ['Total', 'Remarks']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = title_font
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = thin_border()

    row += 1
    # Summary rows
    summary_data = [
        ('Total Received Unit (kWh)', [float(m.net_energy_received) for m in months], float(report.total_received_kwh), green_fill),
        ('Total Utilised Unit (kWh)', [float(m.total_energy_utilised) for m in months], float(report.total_utilised_kwh), None),
        ('Loss Unit (kWh)', [float(m.loss_unit) for m in months], float(report.total_loss_kwh), red_fill),
        ('Monthly Loss Percentage', [round(float(m.monthly_loss_percent) * 100, 4) for m in months], None, None),
        ('Cumulative Loss Percentage', [round(float(m.cumulative_loss_percent) * 100, 4) for m in months], round(float(report.cumulative_loss_percent) * 100, 4), None),
    ]

    for label, values, total, fill in summary_data:
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].alignment = left
        if fill:
            ws[f'A{row}'].fill = fill
        for ci, v in enumerate(values, 5):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = number_font
            cell.alignment = right
            cell.number_format = '#,##0.00'
            cell.border = thin_border()
            if fill:
                cell.fill = fill
        total_cell = ws.cell(row=row, column=5 + len(values), value=total if total is not None else '')
        total_cell.font = Font(name='Arial', bold=True, size=10)
        total_cell.alignment = right
        total_cell.border = thin_border()
        if fill:
            total_cell.fill = fill
        row += 1

    row += 1

    # Preload data for section tables
    import_types = ['SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP']
    export_types = ['EXPORT_DC', 'EXPORT_IPP']
    all_points = MeterPoint.objects.filter(
        distribution_center=report.distribution_center,
        source_type__in=import_types + export_types,
    )

    month_ids = [m.pk for m in months]
    meter_readings = MeterReading.objects.filter(
        monthly_data_id__in=month_ids,
        meter_point__in=all_points,
    ).select_related('meter_point')
    
    # Create detailed reading map for all meter readings
    reading_map = {(r.monthly_data_id, r.meter_point_id): r for r in meter_readings}
    
    # Also create unit_kwh map for summary calculations
    unit_kwh_map = {(r.monthly_data_id, r.meter_point_id): float(r.unit_kwh) for r in meter_readings}

    # Build set of (monthly_data_id, meter_point_id) pairs deleted for that specific month.
    # These cells must show as blank (not 0) — feeder was not present in that month.
    deleted_for_month = set(
        MonthlyMeterPointStatus.objects.filter(
            monthly_data_id__in=month_ids,
            is_active=False,
        ).values_list('monthly_data_id', 'meter_point_id')
    )

    cats = ConsumerCategory.objects.filter(is_active=True).filter(
        Q(distribution_center__isnull=True) | Q(distribution_center_id=report.distribution_center_id)
    ).order_by('display_order', 'name')

    energy_utilisations = EnergyUtilisation.objects.filter(
        monthly_data_id__in=month_ids,
        consumer_category__in=cats,
    )
    eu_map = {(e.monthly_data_id, e.consumer_category_id): e for e in energy_utilisations}

    consumer_counts = ConsumerCount.objects.filter(
        monthly_data_id__in=month_ids,
        consumer_category__in=cats,
    )
    cc_map = {(c.monthly_data_id, c.consumer_category_id): c for c in consumer_counts}

    def write_energy_section(title, points, source_types, start_row):
        points = list(points)
        # Columns: A=S.No, B=Name, then months, then Total
        last_col = 2 + len(months) + 1
        last_col_letter = get_column_letter(last_col)
        ws.merge_cells(f'A{start_row}:{last_col_letter}{start_row}')
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = title_font
        ws[f'A{start_row}'].fill = gray_fill
        ws[f'A{start_row}'].alignment = left

        header_row = start_row + 1
        headers = ['S.No.', 'Meter / Consumer Source'] + month_names + ['Total (kWh)']
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=ci, value=h)
            cell.font = Font(name='Arial', bold=True, size=9)
            cell.fill = blue_fill
            cell.alignment = center
            cell.border = thin_border()

        table_row = header_row + 1
        for idx, mp in enumerate(points, 1):
            ws.cell(row=table_row, column=1, value=idx).alignment = center
            ws.cell(row=table_row, column=2, value=mp.name).alignment = left
            row_sum = 0.0
            for mi, md in enumerate(months):
                # If feeder was deleted for this specific month, leave cell blank
                if (md.pk, mp.pk) in deleted_for_month:
                    cell = ws.cell(row=table_row, column=3 + mi, value='—')
                    cell.alignment = center
                    cell.font = Font(name='Arial', size=9, color='AAAAAA')
                elif mp.source_type in source_types:
                    val = unit_kwh_map.get((md.pk, mp.pk), 0.0)
                    row_sum += val
                    cell = ws.cell(row=table_row, column=3 + mi, value=val)
                    cell.number_format = '#,##0.00'
                    cell.alignment = right
                else:
                    cell = ws.cell(row=table_row, column=3 + mi, value=0.0)
                    cell.number_format = '#,##0.00'
                    cell.alignment = right
            total_cell = ws.cell(row=table_row, column=3 + len(months), value=row_sum)
            total_cell.number_format = '#,##0.00'
            total_cell.font = Font(name='Arial', bold=True, size=9)
            total_cell.alignment = right
            table_row += 1

        # Grand total
        ws.cell(row=table_row, column=2, value='Grand Total:').font = title_font
        grand = 0.0
        for mi, md in enumerate(months):
            col_sum = 0.0
            for mp in points:
                if (md.pk, mp.pk) not in deleted_for_month and mp.source_type in source_types:
                    col_sum += unit_kwh_map.get((md.pk, mp.pk), 0.0)
            grand += col_sum
            cell = ws.cell(row=table_row, column=3 + mi, value=col_sum)
            cell.number_format = '#,##0.00'
            cell.alignment = right
            cell.fill = green_fill
        total_cell = ws.cell(row=table_row, column=3 + len(months), value=grand)
        total_cell.number_format = '#,##0.00'
        total_cell.font = Font(name='Arial', bold=True, size=9)
        total_cell.alignment = right
        return table_row + 1

    # Section A - Energy Import (Detailed Format)
    dc_import_points = all_points.filter(source_type__in=import_types).order_by('source_type', 'name')
    
    if not dc_import_points.exists():
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = 'No Import Meter Points Configured for this Distribution Center'
        ws[f'A{row}'].font = Font(name='Arial', size=11, color='C0392B')
        ws[f'A{row}'].alignment = center
        row += 2
    else:
        # Section A Header
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = 'A. Energy Import - Detailed Meter Readings'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = gray_fill
        ws[f'A{row}'].alignment = left
        row += 1
        
        # Headers for detailed readings
        detailed_headers = ['S.No.', 'Meter Point', 'Month', 'Previous Reading', 'Present Reading', 'Difference', 'Multiplying Factor', 'Unit (kWh)']
        for ci, h in enumerate(detailed_headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = Font(name='Arial', bold=True, size=9)
            cell.fill = blue_fill
            cell.alignment = center
            cell.border = thin_border()
        row += 1
        
        # Detailed meter readings data for Import points
        for idx, mp in enumerate(dc_import_points, 1):
            for mi, md in enumerate(months):
                reading = reading_map.get((md.pk, mp.pk))
                if reading and float(reading.present_reading) > 0:  # Only show rows with actual data
                    ws.cell(row=row, column=1, value=idx).alignment = center
                    ws.cell(row=row, column=2, value=mp.name).alignment = left
                    ws.cell(row=row, column=3, value=md.month_name).alignment = center
                    ws.cell(row=row, column=4, value=float(reading.previous_reading)).number_format = '#,##0.000'
                    ws.cell(row=row, column=5, value=float(reading.present_reading)).number_format = '#,##0.000'
                    ws.cell(row=row, column=6, value=float(reading.difference)).number_format = '#,##0.000'
                    ws.cell(row=row, column=7, value=float(reading.multiplying_factor)).number_format = '#,##0.000'
                    ws.cell(row=row, column=8, value=float(reading.unit_kwh)).number_format = '#,##0.00'
                    
                    # Add borders to all cells
                    for col in range(1, 9):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border()
                        cell.font = Font(name='Arial', size=8)
                    
                    row += 1
            idx += 1

    row += 2

    # Section B - Energy Export (Detailed Format)
    dc_export_points = all_points.filter(source_type__in=export_types).order_by('source_type', 'name')
    
    if not dc_export_points.exists():
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = 'No Export Meter Points Configured for this Distribution Center'
        ws[f'A{row}'].font = Font(name='Arial', size=11, color='C0392B')
        ws[f'A{row}'].alignment = center
        row += 2
    else:
        # Section B Header
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = 'B. Energy Export - Detailed Meter Readings'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = gray_fill
        ws[f'A{row}'].alignment = left
        row += 1
        
        # Headers for detailed readings
        detailed_headers = ['S.No.', 'Meter Point', 'Month', 'Previous Reading', 'Present Reading', 'Difference', 'Multiplying Factor', 'Unit (kWh)']
        for ci, h in enumerate(detailed_headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = Font(name='Arial', bold=True, size=9)
            cell.fill = blue_fill
            cell.alignment = center
            cell.border = thin_border()
        row += 1
        
        # Detailed meter readings data for Export points
        for idx, mp in enumerate(dc_export_points, 1):
            for mi, md in enumerate(months):
                reading = reading_map.get((md.pk, mp.pk))
                if reading and float(reading.present_reading) > 0:  # Only show rows with actual data
                    ws.cell(row=row, column=1, value=idx).alignment = center
                    ws.cell(row=row, column=2, value=mp.name).alignment = left
                    ws.cell(row=row, column=3, value=md.month_name).alignment = center
                    ws.cell(row=row, column=4, value=float(reading.previous_reading)).number_format = '#,##0.000'
                    ws.cell(row=row, column=5, value=float(reading.present_reading)).number_format = '#,##0.000'
                    ws.cell(row=row, column=6, value=float(reading.difference)).number_format = '#,##0.000'
                    ws.cell(row=row, column=7, value=float(reading.multiplying_factor)).number_format = '#,##0.000'
                    ws.cell(row=row, column=8, value=float(reading.unit_kwh)).number_format = '#,##0.00'
                    
                    # Add borders to all cells
                    for col in range(1, 9):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border()
                        cell.font = Font(name='Arial', size=8)
                    
                    row += 1
            idx += 1

    # Section C - Net Received
    last_col = 2 + len(months) + 1
    last_col_letter = get_column_letter(last_col)
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    ws[f'A{row}'] = 'C. Net Energy Received (kWh)'
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = gray_fill
    ws[f'A{row}'].alignment = left
    header_row = row + 1
    headers = ['Particular'] + month_names + ['Total (kWh)']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = Font(name='Arial', bold=True, size=9)
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = thin_border()
    row += 2
    ws.cell(row=row, column=1, value='Net Received').font = title_font
    grand_net = 0.0
    for mi, md in enumerate(months):
        val = float(md.net_energy_received)
        grand_net += val
        cell = ws.cell(row=row, column=2 + mi, value=val)
        cell.number_format = '#,##0.00'
        cell.alignment = right
    total_cell = ws.cell(row=row, column=2 + len(months), value=float(report.total_received_kwh))
    total_cell.number_format = '#,##0.00'
    total_cell.font = Font(name='Arial', bold=True, size=9)
    total_cell.alignment = right
    row += 2

    # Section D - Energy Utilised (consumer categories)
    last_col = 2 + len(months) + 1
    last_col_letter = get_column_letter(last_col)
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    ws[f'A{row}'] = 'D. Energy Utilised (kWh) by Consumer Category'
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = gray_fill
    ws[f'A{row}'].alignment = left
    header_row = row + 1
    headers = ['S.No.', 'Consumer Category'] + month_names + ['Total (kWh)']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = Font(name='Arial', bold=True, size=9)
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = thin_border()
    table_row = header_row + 1
    for idx, cat in enumerate(cats, 1):
        ws.cell(row=table_row, column=1, value=idx).alignment = center
        ws.cell(row=table_row, column=2, value=cat.name).alignment = left
        row_sum = 0.0
        for mi, md in enumerate(months):
            eu = eu_map.get((md.pk, cat.pk))
            val = float(eu.energy_kwh) if eu else 0.0
            row_sum += val
            cell = ws.cell(row=table_row, column=3 + mi, value=val)
            cell.number_format = '#,##0.00'
            cell.alignment = right
        total_cell = ws.cell(row=table_row, column=3 + len(months), value=row_sum)
        total_cell.number_format = '#,##0.00'
        total_cell.font = Font(name='Arial', bold=True, size=9)
        total_cell.alignment = right
        table_row += 1

    # Grand total row
    ws.cell(row=table_row, column=2, value='Grand Total:').font = title_font
    for mi, md in enumerate(months):
        val = float(md.total_energy_utilised)
        cell = ws.cell(row=table_row, column=3 + mi, value=val)
        cell.number_format = '#,##0.00'
        cell.alignment = right
        cell.fill = green_fill
    total_cell = ws.cell(row=table_row, column=3 + len(months), value=float(report.total_utilised_kwh))
    total_cell.number_format = '#,##0.00'
    total_cell.font = Font(name='Arial', bold=True, size=9)
    total_cell.alignment = right
    row = table_row + 2

    # Section E - Consumer Count (consumer categories)
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    ws[f'A{row}'] = 'E. Consumer Count by Consumer Category'
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = gray_fill
    ws[f'A{row}'].alignment = left
    header_row = row + 1
    headers = ['S.No.', 'Consumer Category'] + month_names + ['Total Count']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = Font(name='Arial', bold=True, size=9)
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = thin_border()

    table_row = header_row + 1
    for idx, cat in enumerate(cats, 1):
        ws.cell(row=table_row, column=1, value=idx).alignment = center
        ws.cell(row=table_row, column=2, value=cat.name).alignment = left
        row_sum = 0
        for mi, md in enumerate(months):
            cc = cc_map.get((md.pk, cat.pk))
            val = int(cc.count) if cc else 0
            row_sum += val
            cell = ws.cell(row=table_row, column=3 + mi, value=val)
            cell.alignment = right
        total_cell = ws.cell(row=table_row, column=3 + len(months), value=row_sum)
        total_cell.font = Font(name='Arial', bold=True, size=9)
        total_cell.alignment = right
        table_row += 1

    row = table_row + 1
    return wb